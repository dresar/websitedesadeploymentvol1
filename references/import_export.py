"""
Sistem Import/Export untuk References App
Menyediakan fitur import/export data dengan format Excel, CSV, JSON, dan PDF
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from datetime import datetime, date
import json
import io
import os
from typing import Dict, List, Any, Optional, Tuple
import logging

# PDF support
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import (
    Penduduk, Family, Dusun, Lorong, RW, RT, Pelajar, DisabilitasType, 
    DisabilitasData, ReligionReference, Keluarga
)

logger = logging.getLogger(__name__)

class ImportExportManager:
    """Manager untuk operasi import/export data"""
    
    def __init__(self):
        self.supported_formats = ['excel', 'csv', 'json']
        if PDF_AVAILABLE:
            self.supported_formats.append('pdf')
        
        self.model_mapping = {
            'penduduk': Penduduk,
            'keluarga': Keluarga,
            'family': Family,
            'dusun': Dusun,
            'lorong': Lorong,
            'rw': RW,
            'rt': RT,
            'pelajar': Pelajar,
            'disabilitas': DisabilitasData,
            'disabilitas_type': DisabilitasType,
            'religion': ReligionReference,
        }
        
        # Field mappings for each model to ensure consistent export/import
        self.field_mappings = {
            'penduduk': {
                'nik': 'NIK',
                'name': 'Nama Lengkap',
                'gender': 'Jenis Kelamin',
                'birth_place': 'Tempat Lahir',
                'birth_date': 'Tanggal Lahir',
                'kk_number': 'Nomor KK',
                'religion': 'Agama',
                'education': 'Pendidikan',
                'occupation': 'Pekerjaan',
                'marital_status': 'Status Perkawinan',
                'blood_type': 'Golongan Darah',
                'phone_number': 'Nomor Telepon',
                'mobile_number': 'Nomor HP',
                'email': 'Email',
                'dusun': 'Dusun',
                'lorong': 'Lorong',
                'rw': 'RW',
                'rt': 'RT',
                'rt_number': 'Nomor RT',
                'rw_number': 'Nomor RW',
                'house_number': 'Nomor Rumah',
                'address': 'Alamat',
                'citizenship': 'Kewarganegaraan',
                'is_active': 'Status Aktif',
                'is_alive': 'Status Hidup',
            },
            'dusun': {
                'name': 'Nama Dusun',
                'code': 'Kode Dusun',
                'description': 'Deskripsi',
                'area_size': 'Luas Area (Ha)',
                'population_count': 'Jumlah Penduduk',
                'kepala_dusun': 'Kepala Dusun',
                'is_active': 'Status Aktif',
            },
            'lorong': {
                'dusun': 'Dusun',
                'nama_lorong': 'Nama Lorong',
                'kode': 'Kode Lorong',
                'ketua_lorong': 'Ketua Lorong',
                'rt_number': 'Nomor RT',
                'description': 'Deskripsi',
                'length': 'Panjang (m)',
                'house_count': 'Jumlah Rumah',
                'population_count': 'Jumlah Penduduk',
                'is_active': 'Status Aktif',
            },
            'rw': {
                'dusun': 'Dusun',
                'rw_number': 'Nomor RW',
                'ketua_rw': 'Ketua RW',
                'description': 'Deskripsi',
                'population_count': 'Jumlah Penduduk',
                'is_active': 'Status Aktif',
            },
            'rt': {
                'rw': 'RW',
                'rt_number': 'Nomor RT',
                'ketua_rt': 'Ketua RT',
                'description': 'Deskripsi',
                'population_count': 'Jumlah Penduduk',
                'is_active': 'Status Aktif',
            },
            'keluarga': {
                'dusun': 'Dusun',
                'nama_kepala_keluarga': 'Nama Kepala Keluarga',
                'nomor_kk': 'Nomor KK',
                'rt': 'RT',
                'rw': 'RW',
                'alamat': 'Alamat',
                'is_active': 'Status Aktif',
            },
            'family': {
                'kk_number': 'Nomor KK',
                'head': 'Kepala Keluarga',
                'family_status': 'Status Keluarga',
                'total_members': 'Jumlah Anggota',
                'total_income': 'Total Pendapatan',
                'address': 'Alamat',
                'dusun': 'Dusun',
                'lorong': 'Lorong',
                'rt_number': 'Nomor RT',
                'rw_number': 'Nomor RW',
                'house_number': 'Nomor Rumah',
                'postal_code': 'Kode Pos',
                'phone_number': 'Nomor Telepon',
                'is_active': 'Status Aktif',
            },
            'pelajar': {
                'penduduk': 'Nama Penduduk',
                'jenjang': 'Jenjang Pendidikan',
                'sekolah': 'Nama Sekolah',
                'tahun_masuk': 'Tahun Masuk',
                'status': 'Status',
                'keterangan': 'Keterangan',
                'is_active': 'Status Aktif',
            },
            'disabilitas': {
                'penduduk': 'Nama Penduduk',
                'disability_type': 'Jenis Disabilitas',
                'severity': 'Tingkat Keparahan',
                'description': 'Deskripsi',
                'diagnosis_date': 'Tanggal Diagnosis',
                'needs_assistance': 'Membutuhkan Bantuan',
                'is_active': 'Status Aktif',
            },
            'disabilitas_type': {
                'name': 'Nama Jenis',
                'code': 'Kode',
                'description': 'Deskripsi',
                'is_active': 'Status Aktif',
            },
            'religion': {
                'name': 'Nama Agama',
                'code': 'Kode',
                'is_active': 'Status Aktif',
            },
        }
    
    def export_data(self, model_name: str, format_type: str, 
                   filters: Dict = None, include_related: bool = True) -> HttpResponse:
        """Export data dalam format yang ditentukan"""
        try:
            if model_name not in self.model_mapping:
                raise ValueError(f"Model tidak didukung: {model_name}")
            
            model_class = self.model_mapping[model_name]
            queryset = self._apply_filters(model_class, filters)
            
            if format_type == 'excel':
                return self._export_to_excel(model_name, queryset, include_related)
            elif format_type == 'csv':
                return self._export_to_csv(model_name, queryset, include_related)
            elif format_type == 'json':
                return self._export_to_json(model_name, queryset, include_related)
            elif format_type == 'pdf':
                if not PDF_AVAILABLE:
                    raise ValueError("PDF export tidak tersedia. Install reportlab: pip install reportlab")
                return self._export_to_pdf(model_name, queryset, include_related)
            else:
                raise ValueError(f"Format tidak didukung: {format_type}")
                
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    def import_data(self, model_name: str, file_path: str, 
                   format_type: str, validate_only: bool = False) -> Dict[str, Any]:
        """Import data dari file dengan deteksi kolom otomatis untuk penduduk"""
        try:
            if model_name not in self.model_mapping:
                raise ValueError(f"Model tidak didukung: {model_name}")
            
            model_class = self.model_mapping[model_name]
            
            # Baca data berdasarkan format
            if format_type in ['excel', 'xlsx', 'xls']:
                data = self._read_excel_file(file_path)
            elif format_type == 'csv':
                data = self._read_csv_file(file_path)
            elif format_type == 'json':
                data = self._read_json_file(file_path)
            else:
                raise ValueError(f"Format tidak didukung: {format_type}")
            
            # Untuk penduduk, lakukan deteksi dan mapping kolom otomatis
            if model_name == 'penduduk' and data:
                data = self._detect_and_map_penduduk_columns(data)
            
            # Validasi dan import data
            result = self._import_data_to_model(model_class, data, validate_only)
            return result
            
        except Exception as e:
            logger.error(f"Import error: {str(e)}")
            return {'success': False, 'error': str(e), 'imported': 0, 'errors': []}
    
    def create_template(self, model_name: str, format_type: str = 'excel') -> HttpResponse:
        """Buat template untuk import data dari folder templates_import/excel_templates/"""
        try:
            import os
            from django.conf import settings
            
            if format_type != 'excel':
                # Untuk format lain, gunakan method lama
                return self._create_csv_template(model_name) if format_type == 'csv' else self._create_json_template(model_name)
            
            # Path ke folder templates
            templates_dir = os.path.join(settings.BASE_DIR, 'templates_import', 'excel_templates')
            
            # Mapping nama model ke nama file template
            template_files = {
                'penduduk': 'template_import_penduduk.xlsx',
                'dusun': 'template_import_dusun.xlsx',
                'lorong': 'template_import_lorong.xlsx',
                'rw': 'template_import_rw.xlsx',
                'rt': 'template_import_rt.xlsx',
                'keluarga': 'template_import_keluarga.xlsx',
                'pelajar': 'template_import_pelajar.xlsx',
                'disabilitas': 'template_import_disabilitas.xlsx',
            }
            
            if model_name not in template_files:
                raise ValueError(f"Template untuk {model_name} tidak tersedia")
            
            template_file = template_files[model_name]
            template_path = os.path.join(templates_dir, template_file)
            
            if not os.path.exists(template_path):
                # Fallback ke method lama jika template tidak ditemukan
                logger.warning(f"Template {template_file} tidak ditemukan, menggunakan method lama")
                return self._create_excel_template(model_name)
            
            # Baca file template dan kirim sebagai response
            with open(template_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{template_file}"'
                return response
                
        except Exception as e:
            logger.error(f"Template creation error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    def _apply_filters(self, model_class, filters: Dict = None):
        """Terapkan filter ke queryset"""
        queryset = model_class.objects.all()
        
        if not filters:
            return queryset
        
        # Terapkan filter umum
        if 'is_active' in filters:
            queryset = queryset.filter(is_active=filters['is_active'])
        
        if 'date_from' in filters:
            queryset = queryset.filter(created_at__gte=filters['date_from'])
        
        if 'date_to' in filters:
            queryset = queryset.filter(created_at__lte=filters['date_to'])
        
        # Terapkan filter khusus model
        if model_class == Penduduk:
            if 'dusun' in filters:
                queryset = queryset.filter(dusun__name__icontains=filters['dusun'])
            if 'gender' in filters:
                queryset = queryset.filter(gender=filters['gender'])
            if 'religion' in filters:
                queryset = queryset.filter(religion=filters['religion'])
        
        return queryset
    
    def _queryset_to_dataframe_with_mapping(self, model_name: str, queryset, include_related: bool = True) -> pd.DataFrame:
        """Convert queryset to DataFrame using field mappings"""
        try:
            if model_name not in self.field_mappings:
                raise ValueError(f"No field mapping found for model: {model_name}")
            
            field_mapping = self.field_mappings[model_name]
            data = []
            
            for obj in queryset:
                row = {}
                for field_name, display_name in field_mapping.items():
                    value = self._get_field_value_with_mapping(obj, field_name, include_related)
                    row[display_name] = value
                data.append(row)
            
            return pd.DataFrame(data)
            
        except Exception as e:
            logger.error(f"DataFrame conversion error: {str(e)}")
            raise
    
    def _get_field_value_with_mapping(self, obj, field_name: str, include_related: bool = True):
        """Get field value with proper mapping and formatting"""
        try:
            if '.' in field_name:
                # Handle related field access
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        return None
            else:
                value = getattr(obj, field_name, None)
            
            # Format specific field types
            if value is None:
                return ''
            elif isinstance(value, (datetime, date)):
                return value.strftime('%Y-%m-%d')
            elif isinstance(value, bool):
                return 'Ya' if value else 'Tidak'
            elif hasattr(value, 'name'):  # Foreign key with name attribute
                return value.name
            elif hasattr(value, '__str__'):
                return str(value)
            else:
                return value
                
        except Exception as e:
            logger.error(f"Field value error for {field_name}: {str(e)}")
            return ''
    
    def _export_to_excel(self, model_name: str, queryset, include_related: bool = True) -> HttpResponse:
        """Export data ke Excel dengan formatting menggunakan field mapping"""
        try:
            # Convert queryset to DataFrame with field mapping
            df = self._queryset_to_dataframe_with_mapping(model_name, queryset, include_related)
            
            # Create Excel file in memory
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=f"Data {model_name.title()}", index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[f"Data {model_name.title()}"]
                
                # Apply comprehensive styling
                self._apply_comprehensive_excel_styling(worksheet, df)
            
            output.seek(0)
            
            # Create HTTP response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{model_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            raise
    
    def _apply_comprehensive_excel_styling(self, worksheet, df):
        """Apply comprehensive styling to Excel worksheet"""
        try:
            # Style header row
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header styling
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # Style data rows
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal="left", vertical="center")
            data_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply data styling
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border
            
            # Auto-adjust column widths
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Check header length
                header_length = len(str(df.columns[col-1]))
                max_length = max(max_length, header_length)
                
                # Check data lengths
                for row in range(2, len(df) + 2):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            logger.error(f"Excel styling error: {str(e)}")
            # Continue without styling if there's an error
    
    def _export_to_csv(self, model_name: str, queryset, include_related: bool = True) -> HttpResponse:
        """Export data ke CSV menggunakan field mapping"""
        try:
            # Convert queryset to DataFrame with field mapping
            df = self._queryset_to_dataframe_with_mapping(model_name, queryset, include_related)
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{model_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            # Write CSV with proper encoding
            df.to_csv(response, index=False, encoding='utf-8-sig')
            
            return response
            
        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            raise
    
    def _export_to_json(self, model_name: str, queryset, include_related: bool = True) -> HttpResponse:
        """Export data ke JSON menggunakan field mapping"""
        try:
            # Convert queryset to DataFrame with field mapping
            df = self._queryset_to_dataframe_with_mapping(model_name, queryset, include_related)
            
            # Convert DataFrame to JSON
            data = df.to_dict('records')
            
            response = HttpResponse(
                json.dumps(data, indent=2, ensure_ascii=False),
                content_type='application/json; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename="{model_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
            
            return response
            
        except Exception as e:
            logger.error(f"JSON export error: {str(e)}")
            raise
    
    def _create_excel_template(self, model_name: str) -> HttpResponse:
        """Buat template Excel untuk import dengan data contoh lengkap"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Template {model_name.title()}"
        
        # Dapatkan konfigurasi field
        field_config = self._get_field_config(model_name)
        
        # Tulis header dengan styling yang lebih baik
        headers = [config['verbose_name'] for config in field_config]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color="FFFFFF"),
                right=Side(style='thin', color="FFFFFF"),
                top=Side(style='thin', color="FFFFFF"),
                bottom=Side(style='thin', color="FFFFFF")
            )
        
        # Tambahkan contoh data dengan styling
        example_data = self._get_example_data(model_name)
        for row_idx, example in enumerate(example_data, 2):
            for col_idx, config in enumerate(field_config, 1):
                field_name = config['field_name']
                value = example.get(field_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Styling untuk data
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color="DEE2E6"),
                    right=Side(style='thin', color="DEE2E6"),
                    top=Side(style='thin', color="DEE2E6"),
                    bottom=Side(style='thin', color="DEE2E6")
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Tambahkan sheet instruksi
        instruction_ws = wb.create_sheet("Instruksi", 0)
        instruction_ws.title = "Instruksi"
        
        # Instruksi berdasarkan model
        instructions = self._get_import_instructions(model_name)
        for idx, instruction in enumerate(instructions, 1):
            instruction_ws.cell(row=idx, column=1, value=instruction)
            instruction_ws.cell(row=idx, column=1).font = Font(size=11)
        
        # Auto-adjust column widths
        for ws_sheet in wb.worksheets:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Buat response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="template_{model_name}_lengkap.xlsx"'
        
        wb.save(response)
        return response
    
    def _create_csv_template(self, model_name: str) -> HttpResponse:
        """Buat template CSV untuk import"""
        field_config = self._get_field_config(model_name)
        example_data = self._get_example_data(model_name)
        
        # Siapkan data
        data = []
        for example in example_data:
            row = {}
            for config in field_config:
                field_name = config['field_name']
                row[config['verbose_name']] = example.get(field_name, '')
            data.append(row)
        
        # Buat DataFrame dan export
        df = pd.DataFrame(data)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="template_{model_name}.csv"'
        
        df.to_csv(response, index=False, encoding='utf-8')
        return response
    
    def _create_json_template(self, model_name: str) -> HttpResponse:
        """Buat template JSON untuk import"""
        example_data = self._get_example_data(model_name)
        
        response = HttpResponse(
            json.dumps(example_data, indent=2, ensure_ascii=False),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="template_{model_name}.json"'
        
        return response
    
    def _get_field_config(self, model_name: str) -> List[Dict[str, Any]]:
        """Dapatkan konfigurasi field untuk model"""
        configs = {
            'penduduk': [
                {'field_name': 'nik', 'verbose_name': 'NIK'},
                {'field_name': 'name', 'verbose_name': 'Nama Lengkap'},
                {'field_name': 'gender', 'verbose_name': 'Jenis Kelamin'},
                {'field_name': 'birth_place', 'verbose_name': 'Tempat Lahir'},
                {'field_name': 'birth_date', 'verbose_name': 'Tanggal Lahir'},
                {'field_name': 'kk_number', 'verbose_name': 'Nomor KK'},
                {'field_name': 'religion', 'verbose_name': 'Agama'},
                {'field_name': 'education', 'verbose_name': 'Pendidikan'},
                {'field_name': 'occupation', 'verbose_name': 'Pekerjaan'},
                {'field_name': 'marital_status', 'verbose_name': 'Status Perkawinan'},
                {'field_name': 'dusun', 'verbose_name': 'Dusun'},
                {'field_name': 'lorong', 'verbose_name': 'Lorong'},
                {'field_name': 'rt', 'verbose_name': 'RT'},
                {'field_name': 'rw', 'verbose_name': 'RW'},
                {'field_name': 'address', 'verbose_name': 'Alamat'},
                {'field_name': 'phone_number', 'verbose_name': 'Telepon'},
                {'field_name': 'relationship_to_head', 'verbose_name': 'Hubungan dengan Kepala Keluarga'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'dusun': [
                {'field_name': 'name', 'verbose_name': 'Nama Dusun'},
                {'field_name': 'code', 'verbose_name': 'Kode Dusun'},
                {'field_name': 'description', 'verbose_name': 'Deskripsi'},
                {'field_name': 'area_size', 'verbose_name': 'Luas Area (Ha)'},
                {'field_name': 'alamat_kantor', 'verbose_name': 'Alamat Kantor'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'keluarga': [
                {'field_name': 'nomor_kk', 'verbose_name': 'Nomor KK'},
                {'field_name': 'nama_kepala_keluarga', 'verbose_name': 'Nama Kepala Keluarga'},
                {'field_name': 'alamat', 'verbose_name': 'Alamat'},
                {'field_name': 'dusun', 'verbose_name': 'Dusun'},
                {'field_name': 'rt', 'verbose_name': 'RT'},
                {'field_name': 'rw', 'verbose_name': 'RW'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'pelajar': [
                {'field_name': 'penduduk', 'verbose_name': 'Nama Penduduk'},
                {'field_name': 'jenjang', 'verbose_name': 'Jenjang Pendidikan'},
                {'field_name': 'sekolah', 'verbose_name': 'Nama Sekolah'},
                {'field_name': 'status', 'verbose_name': 'Status'},
                {'field_name': 'tahun_masuk', 'verbose_name': 'Tahun Masuk'},
                {'field_name': 'keterangan', 'verbose_name': 'Keterangan'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'lorong': [
                {'field_name': 'nama_lorong', 'verbose_name': 'Nama Lorong'},
                {'field_name': 'kode', 'verbose_name': 'Kode Lorong'},
                {'field_name': 'dusun', 'verbose_name': 'Dusun'},
                {'field_name': 'rt_number', 'verbose_name': 'Nomor RT'},
                {'field_name': 'description', 'verbose_name': 'Deskripsi'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'rw': [
                {'field_name': 'rw_number', 'verbose_name': 'Nomor RW'},
                {'field_name': 'dusun', 'verbose_name': 'Dusun'},
                {'field_name': 'ketua_rw', 'verbose_name': 'Ketua RW'},
                {'field_name': 'description', 'verbose_name': 'Deskripsi'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'rt': [
                {'field_name': 'rt_number', 'verbose_name': 'Nomor RT'},
                {'field_name': 'dusun', 'verbose_name': 'Dusun'},
                {'field_name': 'rw', 'verbose_name': 'RW'},
                {'field_name': 'ketua_rt', 'verbose_name': 'Ketua RT'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
            'disabilitas': [
                {'field_name': 'penduduk', 'verbose_name': 'Nama Penduduk'},
                {'field_name': 'jenis_disabilitas', 'verbose_name': 'Jenis Disabilitas'},
                {'field_name': 'tingkat_disabilitas', 'verbose_name': 'Tingkat Disabilitas'},
                {'field_name': 'keterangan', 'verbose_name': 'Keterangan'},
                {'field_name': 'is_active', 'verbose_name': 'Aktif'},
            ],
        }
        
        return configs.get(model_name, [])
    
    def _get_example_data(self, model_name: str) -> List[Dict[str, Any]]:
        """Dapatkan data contoh lengkap untuk template"""
        examples = {
            'penduduk': [
                {
                    'nik': '1234567890123456',
                    'name': 'Ahmad Suryadi',
                    'gender': 'L',
                    'birth_place': 'Jakarta',
                    'birth_date': '1990-01-15',
                    'kk_number': '1234567890123456',
                    'religion': 'Islam',
                    'education': 'SMA',
                    'occupation': 'Karyawan Swasta',
                    'marital_status': 'KAWIN',
                    'dusun': 'Pulosarok Tengah',
                    'lorong': 'Lorong Masjid',
                    'rt': '001',
                    'rw': '001',
                    'address': 'Jl. Pulosarok Tengah No. 123',
                    'phone_number': '081234567890',
                    'relationship_to_head': 'KEPALA_KELUARGA',
                    'is_active': True,
                },
                {
                    'nik': '1234567890123457',
                    'name': 'Siti Aminah',
                    'gender': 'P',
                    'birth_place': 'Bandung',
                    'birth_date': '1992-05-20',
                    'kk_number': '1234567890123456',
                    'religion': 'Islam',
                    'education': 'SMA',
                    'occupation': 'Ibu Rumah Tangga',
                    'marital_status': 'KAWIN',
                    'dusun': 'Pulosarok Tengah',
                    'lorong': 'Lorong Masjid',
                    'rt': '001',
                    'rw': '001',
                    'address': 'Jl. Pulosarok Tengah No. 123',
                    'phone_number': '081234567891',
                    'relationship_to_head': 'ISTRI',
                    'is_active': True,
                },
                {
                    'nik': '1234567890123458',
                    'name': 'Muhammad Rizki',
                    'gender': 'L',
                    'birth_place': 'Jakarta',
                    'birth_date': '2015-08-10',
                    'kk_number': '1234567890123456',
                    'religion': 'Islam',
                    'education': 'SD',
                    'occupation': 'Pelajar',
                    'marital_status': 'BELUM_KAWIN',
                    'dusun': 'Pulosarok Tengah',
                    'lorong': 'Lorong Masjid',
                    'rt': '001',
                    'rw': '001',
                    'address': 'Jl. Pulosarok Tengah No. 123',
                    'phone_number': '',
                    'relationship_to_head': 'ANAK',
                    'is_active': True,
                }
            ],
            'dusun': [
                {
                    'name': 'Pulosarok Tengah',
                    'code': 'DUSUN-001',
                    'description': 'Dusun utama di tengah desa',
                    'area_size': 25.5,
                    'alamat_kantor': 'Jl. Pulosarok Tengah No. 1',
                    'is_active': True,
                },
                {
                    'name': 'Pulosarok Utara',
                    'code': 'DUSUN-002',
                    'description': 'Dusun di bagian utara desa',
                    'area_size': 30.2,
                    'alamat_kantor': 'Jl. Pulosarok Utara No. 1',
                    'is_active': True,
                },
                {
                    'name': 'Pulosarok Selatan',
                    'code': 'DUSUN-003',
                    'description': 'Dusun di bagian selatan desa',
                    'area_size': 28.8,
                    'alamat_kantor': 'Jl. Pulosarok Selatan No. 1',
                    'is_active': True,
                }
            ],
            'keluarga': [
                {
                    'nomor_kk': '1234567890123456',
                    'nama_kepala_keluarga': 'Ahmad Suryadi',
                    'alamat': 'Jl. Pulosarok Tengah No. 123',
                    'dusun': 'Pulosarok Tengah',
                    'rt': '001',
                    'rw': '001',
                    'is_active': True,
                },
                {
                    'nomor_kk': '1234567890123457',
                    'nama_kepala_keluarga': 'Budi Santoso',
                    'alamat': 'Jl. Pulosarok Utara No. 45',
                    'dusun': 'Pulosarok Utara',
                    'rt': '002',
                    'rw': '001',
                    'is_active': True,
                }
            ],
            'pelajar': [
                {
                    'penduduk': 'Ahmad Suryadi',
                    'jenjang': 'SMA',
                    'sekolah': 'SMA Negeri 1 Pulosarok',
                    'status': 'AKTIF',
                    'tahun_masuk': 2020,
                    'keterangan': 'Siswa berprestasi',
                    'is_active': True,
                },
                {
                    'penduduk': 'Muhammad Rizki',
                    'jenjang': 'SD',
                    'sekolah': 'SD Negeri 1 Pulosarok',
                    'status': 'AKTIF',
                    'tahun_masuk': 2021,
                    'keterangan': 'Siswa kelas 3',
                    'is_active': True,
                }
            ],
            'lorong': [
                {
                    'nama_lorong': 'Lorong Masjid',
                    'kode': 'LORONG-001',
                    'dusun': 'Pulosarok Tengah',
                    'rt_number': '001',
                    'description': 'Lorong di dekat masjid',
                    'is_active': True,
                },
                {
                    'nama_lorong': 'Lorong Pasar',
                    'kode': 'LORONG-002',
                    'dusun': 'Pulosarok Tengah',
                    'rt_number': '002',
                    'description': 'Lorong di dekat pasar',
                    'is_active': True,
                }
            ],
            'rw': [
                {
                    'rw_number': '001',
                    'dusun': 'Pulosarok Tengah',
                    'ketua_rw': 'Ahmad Suryadi',
                    'description': 'RW 001 Pulosarok Tengah',
                    'is_active': True,
                },
                {
                    'rw_number': '002',
                    'dusun': 'Pulosarok Tengah',
                    'ketua_rw': 'Budi Santoso',
                    'description': 'RW 002 Pulosarok Tengah',
                    'is_active': True,
                }
            ],
            'rt': [
                {
                    'rt_number': '001',
                    'dusun': 'Pulosarok Tengah',
                    'rw': '001',
                    'ketua_rt': 'Ahmad Suryadi',
                    'is_active': True,
                },
                {
                    'rt_number': '002',
                    'dusun': 'Pulosarok Tengah',
                    'rw': '001',
                    'ketua_rt': 'Budi Santoso',
                    'is_active': True,
                }
            ],
            'disabilitas': [
                {
                    'penduduk': 'Ahmad Suryadi',
                    'jenis_disabilitas': 'TUNANETRA',
                    'tingkat_disabilitas': 'RINGAN',
                    'keterangan': 'Menggunakan kacamata',
                    'is_active': True,
                },
                {
                    'penduduk': 'Siti Aminah',
                    'jenis_disabilitas': 'TUNARUNGU',
                    'tingkat_disabilitas': 'SEDANG',
                    'keterangan': 'Menggunakan alat bantu dengar',
                    'is_active': True,
                }
            ],
        }
        
        return examples.get(model_name, [])
    
    def _get_import_instructions(self, model_name: str) -> List[str]:
        """Dapatkan instruksi import untuk setiap model"""
        instructions = {
            'penduduk': [
                "📋 INSTRUKSI IMPORT DATA PENDUDUK",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Isi data sesuai dengan contoh yang sudah ada",
                "   • NIK harus unik (16 digit)",
                "   • Nama lengkap sesuai KTP",
                "   • Jenis Kelamin: L (Laki-laki) atau P (Perempuan)",
                "   • Tanggal Lahir format: YYYY-MM-DD (contoh: 1990-01-15)",
                "   • Status Perkawinan: BELUM_KAWIN, KAWIN, CERAI_HIDUP, CERAI_MATI",
                "   • Hubungan dengan Kepala Keluarga: KEPALA_KELUARGA, ISTRI, ANAK, dll",
                "",
                "2. 🏘️ WILAYAH ADMINISTRATIF:",
                "   • Dusun: Nama dusun (akan dibuat otomatis jika belum ada)",
                "   • Lorong: Nama lorong (akan dibuat otomatis jika belum ada)",
                "   • RT: Nomor RT (3 digit, contoh: 001)",
                "   • RW: Nomor RW (3 digit, contoh: 001)",
                "",
                "3. ⚠️ PENTING:",
                "   • Sistem akan otomatis membuat dusun, lorong, RT, RW baru jika belum ada",
                "   • Data duplikat akan dilewati",
                "   • Pastikan NIK unik untuk setiap penduduk",
                "   • Kolom 'Aktif' isi dengan: Ya/True atau Tidak/False",
                "",
                "4. 📊 CONTOH DATA:",
                "   • Lihat baris 2-4 untuk contoh data yang benar",
                "   • Jangan hapus header (baris pertama)",
                "   • Isi data di baris baru sesuai kebutuhan"
            ],
            'dusun': [
                "📋 INSTRUKSI IMPORT DATA DUSUN",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nama Dusun: Nama lengkap dusun",
                "   • Kode Dusun: Kode unik dusun (contoh: DUSUN-001)",
                "   • Deskripsi: Penjelasan tentang dusun",
                "   • Luas Area: Luas wilayah dalam hektar (angka)",
                "   • Alamat Kantor: Alamat kantor dusun",
                "",
                "2. ⚠️ PENTING:",
                "   • Kode dusun harus unik",
                "   • Luas area dalam hektar (gunakan titik untuk desimal)",
                "   • Kolom 'Aktif' isi dengan: Ya/True atau Tidak/False",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-4 untuk contoh data yang benar"
            ],
            'keluarga': [
                "📋 INSTRUKSI IMPORT DATA KELUARGA",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nomor KK: Nomor Kartu Keluarga (16 digit)",
                "   • Nama Kepala Keluarga: Nama lengkap kepala keluarga",
                "   • Alamat: Alamat lengkap keluarga",
                "   • Dusun: Nama dusun (harus sudah ada di database)",
                "   • RT: Nomor RT (3 digit)",
                "   • RW: Nomor RW (3 digit)",
                "",
                "2. ⚠️ PENTING:",
                "   • Nomor KK harus unik",
                "   • Dusun harus sudah ada di database",
                "   • RT dan RW harus sesuai dengan dusun",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ],
            'pelajar': [
                "📋 INSTRUKSI IMPORT DATA PELAJAR",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nama Penduduk: Nama lengkap penduduk (harus sudah ada)",
                "   • Jenjang Pendidikan: SD, SMP, SMA, S1, S2, S3",
                "   • Nama Sekolah: Nama sekolah/universitas",
                "   • Status: AKTIF, LULUS, DROPOUT",
                "   • Tahun Masuk: Tahun masuk sekolah",
                "   • Keterangan: Informasi tambahan",
                "",
                "2. ⚠️ PENTING:",
                "   • Nama penduduk harus sudah ada di database",
                "   • Status: AKTIF, LULUS, atau DROPOUT",
                "   • Tahun masuk dalam format 4 digit (contoh: 2020)",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ],
            'lorong': [
                "📋 INSTRUKSI IMPORT DATA LORONG",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nama Lorong: Nama lengkap lorong",
                "   • Kode Lorong: Kode unik lorong",
                "   • Dusun: Nama dusun (harus sudah ada)",
                "   • Nomor RT: Nomor RT yang dilayani",
                "   • Deskripsi: Penjelasan tentang lorong",
                "",
                "2. ⚠️ PENTING:",
                "   • Kode lorong harus unik",
                "   • Dusun harus sudah ada di database",
                "   • Nomor RT dalam format 3 digit",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ],
            'rw': [
                "📋 INSTRUKSI IMPORT DATA RW",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nomor RW: Nomor RW (3 digit)",
                "   • Dusun: Nama dusun (harus sudah ada)",
                "   • Ketua RW: Nama ketua RW",
                "   • Deskripsi: Penjelasan tentang RW",
                "",
                "2. ⚠️ PENTING:",
                "   • Nomor RW harus unik dalam satu dusun",
                "   • Dusun harus sudah ada di database",
                "   • Nomor RW dalam format 3 digit",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ],
            'rt': [
                "📋 INSTRUKSI IMPORT DATA RT",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nomor RT: Nomor RT (3 digit)",
                "   • Dusun: Nama dusun (harus sudah ada)",
                "   • RW: Nomor RW (harus sudah ada)",
                "   • Ketua RT: Nama ketua RT",
                "",
                "2. ⚠️ PENTING:",
                "   • Nomor RT harus unik dalam satu RW",
                "   • Dusun dan RW harus sudah ada di database",
                "   • Nomor RT dalam format 3 digit",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ],
            'disabilitas': [
                "📋 INSTRUKSI IMPORT DATA DISABILITAS",
                "",
                "1. 📝 CARA MENGISI DATA:",
                "   • Nama Penduduk: Nama lengkap penduduk (harus sudah ada)",
                "   • Jenis Disabilitas: TUNANETRA, TUNARUNGU, TUNANETRA, dll",
                "   • Tingkat Disabilitas: RINGAN, SEDANG, BERAT",
                "   • Keterangan: Penjelasan kondisi disabilitas",
                "",
                "2. ⚠️ PENTING:",
                "   • Nama penduduk harus sudah ada di database",
                "   • Jenis disabilitas sesuai dengan pilihan yang tersedia",
                "   • Tingkat disabilitas: RINGAN, SEDANG, atau BERAT",
                "",
                "3. 📊 CONTOH DATA:",
                "   • Lihat baris 2-3 untuk contoh data yang benar"
            ]
        }
        
        return instructions.get(model_name, ["Instruksi tidak tersedia untuk model ini."])
    
    def _get_field_value(self, obj, field_name: str, include_related: bool = True) -> Any:
        """Dapatkan nilai field dari object"""
        try:
            if '.' in field_name:
                # Handle related fields
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        return None
                return str(value) if value else None
            else:
                value = getattr(obj, field_name, None)
                if hasattr(value, 'name'):
                    return value.name
                elif hasattr(value, 'get_display'):
                    return value.get_display()
                return value
        except:
            return None
    
    def _read_excel_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Baca file Excel dan return data sebagai list of dictionaries"""
        df = pd.read_excel(file_path)
        return df.to_dict('records')
    
    def _read_csv_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Baca file CSV dan return data sebagai list of dictionaries"""
        df = pd.read_csv(file_path, encoding='utf-8')
        return df.to_dict('records')
    
    def _read_json_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Baca file JSON dan return data sebagai list of dictionaries"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    @transaction.atomic
    def _import_data_to_model(self, model_class, data: List[Dict], validate_only: bool = False) -> Dict[str, Any]:
        """Import data ke model dengan validasi dan tracking auto-created items"""
        
        imported = 0
        errors = []
        auto_created = {
            'dusun': 0,
            'lorong': 0,
            'rw': 0,
            'rt': 0
        }
        
        for row_idx, row_data in enumerate(data, 1):
            try:
                print(f"Processing row {row_idx}: {row_data}")
                # Bersihkan dan validasi data
                cleaned_data = self._clean_import_data(model_class, row_data)
                
                # Special handling for RT model
                if model_class.__name__ == 'RT':
                    # Ensure RW field is present and valid
                    if cleaned_data is None or 'rw' not in cleaned_data or cleaned_data['rw'] is None:
                        print(f"SKIP row {row_idx}: RT requires valid RW field")
                        continue
                
                if not validate_only:
                    # Buat atau update object berdasarkan NIK untuk Penduduk
                    if model_class.__name__ == 'Penduduk':
                        # Untuk Penduduk, gunakan NIK sebagai unique identifier
                        nik = cleaned_data.get('nik')
                        if nik:
                            obj, created = model_class.objects.get_or_create(
                                nik=nik,
                                defaults=cleaned_data
                            )
                            # Ensure penduduk is always active
                            if not created:
                                obj.is_active = True
                                obj.is_alive = True
                                obj.save()
                        else:
                            # Jika tidak ada NIK, buat baru
                            obj = model_class.objects.create(**cleaned_data)
                            created = True
                    else:
                        # Untuk model lain, gunakan semua field
                        obj, created = model_class.objects.get_or_create(
                            **cleaned_data
                        )
                    
                    if created:
                        imported += 1
                        
                        # Track auto-created items untuk penduduk
                        if model_class.__name__ == 'Penduduk':
                            if 'dusun' in cleaned_data and hasattr(cleaned_data['dusun'], '_auto_created'):
                                if getattr(cleaned_data['dusun'], '_auto_created', False):
                                    auto_created['dusun'] += 1
                            if 'lorong' in cleaned_data and hasattr(cleaned_data['lorong'], '_auto_created'):
                                if getattr(cleaned_data['lorong'], '_auto_created', False):
                                    auto_created['lorong'] += 1
                            if 'rw' in cleaned_data and hasattr(cleaned_data['rw'], '_auto_created'):
                                if getattr(cleaned_data['rw'], '_auto_created', False):
                                    auto_created['rw'] += 1
                            if 'rt' in cleaned_data and hasattr(cleaned_data['rt'], '_auto_created'):
                                if getattr(cleaned_data['rt'], '_auto_created', False):
                                    auto_created['rt'] += 1
                else:
                    # Hanya validasi
                    obj = model_class(**cleaned_data)
                    obj.full_clean()
                    
            except Exception as e:
                print(f"ERROR in row {row_idx}: {str(e)}")
                print(f"Row data: {row_data}")
                print(f"Exception type: {type(e)}")
                import traceback
                traceback.print_exc()
                errors.append({
                    'row': row_idx,
                    'error': str(e),
                    'data': row_data
                })
        
        # Final validation for Penduduk - ensure all are active
        if model_class.__name__ == 'Penduduk':
            print(f"\n=== FINAL VALIDATION FOR PENDUDUK ===")
            total_penduduk = model_class.objects.count()
            active_penduduk = model_class.objects.filter(is_active=True).count()
            alive_penduduk = model_class.objects.filter(is_alive=True).count()
            
            print(f"Total penduduk: {total_penduduk}")
            print(f"Active penduduk: {active_penduduk}")
            print(f"Alive penduduk: {alive_penduduk}")
            
            # Fix any inactive penduduks
            if active_penduduk < total_penduduk:
                print("Fixing inactive penduduks...")
                fixed_count = model_class.objects.filter(is_active=False).update(is_active=True, is_alive=True)
                print(f"Fixed {fixed_count} penduduks")
                active_penduduk = model_class.objects.filter(is_active=True).count()
                print(f"After fix - Active: {active_penduduk}")
        
        result = {
            'success': len(errors) == 0,
            'imported': imported,
            'errors': errors,
            'total_rows': len(data)
        }
        
        # Tambahkan info auto-created jika ada
        if any(count > 0 for count in auto_created.values()):
            result['auto_created'] = auto_created
        
        return result
    
    def global_import_data(self, file_path: str, format_type: str, validate_only: bool = False) -> Dict[str, Any]:
        """Global import method untuk semua jenis data dalam satu file"""
        try:
            print(f"Format type: {format_type}")
            print(f"Validate only: {validate_only}")
            
            # Baca data dari file
            if format_type in ['excel', 'xlsx', 'xls']:
                print("Reading Excel file...")
                data = self._read_excel_file(file_path)
            elif format_type == 'csv':
                print("Reading CSV file...")
                data = self._read_csv_file(file_path)
            elif format_type == 'json':
                print("Reading JSON file...")
                data = self._read_json_file(file_path)
            else:
                raise ValueError(f"Format tidak didukung: {format_type}")
            
            print(f"Data loaded: {len(data) if data else 0} rows")
            
            if not data:
                return {
                    'success': False,
                    'error': 'File kosong atau tidak dapat dibaca',
                    'imported': 0,
                    'errors': [],
                    'total_rows': 0
                }
            
            # Deteksi model berdasarkan kolom yang ada
            print("Detecting models from data...")
            detected_models = self._detect_models_from_data(data)
            print(f"Detected models: {detected_models}")
            
            if not detected_models:
                return {
                    'success': False,
                    'error': 'Tidak dapat mendeteksi jenis data dari file',
                    'imported': 0,
                    'errors': [],
                    'total_rows': len(data)
                }
            
            # Import untuk setiap model yang terdeteksi
            results = {}
            total_imported = 0
            total_errors = []
            
            for model_name in detected_models:
                try:
                    print(f"Processing model: {model_name}")
                    model_class = self.model_mapping[model_name]
                    model_data = self._filter_data_for_model(data, model_name)
                    
                    if model_data:
                        print(f"Importing {len(model_data)} rows for {model_name}")
                        result = self._import_data_to_model(model_class, model_data, validate_only)
                        results[model_name] = result
                        total_imported += result['imported']
                        total_errors.extend(result['errors'])
                        
                except Exception as e:
                    results[model_name] = {
                        'success': False,
                        'error': str(e),
                        'imported': 0,
                        'errors': [{'row': 0, 'error': str(e)}]
                    }
                    total_errors.append({'row': 0, 'error': f"Error importing {model_name}: {str(e)}"})
            
            return {
                'success': len(total_errors) == 0,
                'imported': total_imported,
                'errors': total_errors,
                'total_rows': len(data),
                'detected_models': detected_models,
                'results': results
            }
            
        except Exception as e:
            logger.error(f"Global import error: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'errors': [{'row': 0, 'error': str(e)}],
                'total_rows': 0
            }
    
    def _determine_model_from_sheet_name(self, sheet_name: str) -> str:
        """Determine model name from sheet name"""
        sheet_name_lower = sheet_name.lower()
        
        # Mapping sheet names to model names
        sheet_to_model = {
            'penduduk': 'penduduk',
            'data penduduk': 'penduduk',
            'penduduk_data': 'penduduk',
            'dusun': 'dusun',
            'data dusun': 'dusun',
            'dusun_data': 'dusun',
            'lorong': 'lorong',
            'data lorong': 'lorong',
            'lorong_data': 'lorong',
            'rw': 'rw',
            'data rw': 'rw',
            'rw_data': 'rw',
            'rt': 'rt',
            'data rt': 'rt',
            'rt_data': 'rt',
            'keluarga': 'keluarga',
            'data keluarga': 'keluarga',
            'keluarga_data': 'keluarga',
            'pelajar': 'pelajar',
            'data pelajar': 'pelajar',
            'pelajar_data': 'pelajar',
            'disabilitas': 'disabilitas',
            'data disabilitas': 'disabilitas',
            'disabilitas_data': 'disabilitas',
        }
        
        return sheet_to_model.get(sheet_name_lower, None)
    
    def _detect_models_from_data(self, data: List[Dict]) -> List[str]:
        """Deteksi model berdasarkan kolom yang ada dalam data"""
        if not data:
            return []
        
        # Ambil kolom dari baris pertama
        columns = list(data[0].keys())
        detected_models = []
        
        # Mapping kolom ke model
        column_model_mapping = {
            'penduduk': ['nik', 'nama lengkap', 'jenis kelamin', 'tempat lahir', 'tanggal lahir'],
            'dusun': ['nama dusun', 'kode dusun', 'luas area'],
            'lorong': ['nama lorong', 'kode lorong', 'dusun'],
            'rw': ['nomor rw', 'rw', 'dusun'],
            'rt': ['nomor rt', 'rt', 'rw'],
            'keluarga': ['nomor kk', 'nama kepala keluarga', 'alamat'],
            'family': ['nomor kk', 'kepala keluarga', 'status keluarga'],
            'pelajar': ['nama penduduk', 'jenjang pendidikan', 'nama sekolah'],
            'disabilitas': ['nama penduduk', 'jenis disabilitas', 'tingkat keparahan'],
            'disabilitas_type': ['nama jenis', 'kode'],
            'religion': ['nama agama', 'kode']
        }
        
        # Normalize kolom names untuk matching
        normalized_columns = [col.lower().strip() for col in columns]
        
        for model_name, model_columns in column_model_mapping.items():
            # Check jika minimal 3 kolom dari model ada dalam data
            matches = sum(1 for col in model_columns if col in normalized_columns)
            if matches >= min(3, len(model_columns)):
                detected_models.append(model_name)
        
        return detected_models
    
    def _filter_data_for_model(self, data: List[Dict], model_name: str) -> List[Dict]:
        """Filter data untuk model tertentu berdasarkan field mapping"""
        if model_name not in self.field_mappings:
            return []
        
        field_mapping = self.field_mappings[model_name]
        filtered_data = []
        
        for row in data:
            filtered_row = {}
            for field_name, display_name in field_mapping.items():
                # Cari kolom yang cocok (case insensitive)
                for col_name, value in row.items():
                    if col_name.lower().strip() == display_name.lower().strip():
                        filtered_row[field_name] = value
                        break
            
            # Hanya tambahkan jika ada minimal 2 field yang cocok
            if len(filtered_row) >= 2:
                filtered_data.append(filtered_row)
        
        return filtered_data
    
    def _clean_import_data(self, model_class, data: Dict[str, Any]) -> Dict[str, Any]:
        """Bersihkan dan validasi data import dengan auto-create untuk penduduk"""
        print(f"Data: {data}")
        
        cleaned_data = {}
        
        # Dapatkan field model
        model_fields = {field.name: field for field in model_class._meta.fields}
        print(f"Model fields: {list(model_fields.keys())}")
        
        # Handle special field mappings for Penduduk dengan auto-create
        if model_class.__name__ == 'Penduduk':
            # Set default values for penduduk
            cleaned_data['is_active'] = True
            cleaned_data['is_alive'] = True
            # Auto-create dusun jika tidak ada
            dusun = None
            if 'dusun' in data and data['dusun']:
                # Import Dusun from references.models
                try:
                    from references.models import Dusun
                except ImportError:
                    Dusun = None
                
                if Dusun is None:
                    print("WARNING: Model Dusun tidak tersedia, skip dusun")
                else:
                    try:
                        dusun = Dusun.objects.get(name__iexact=data['dusun'])
                    except Dusun.DoesNotExist:
                        # Auto-create dusun baru dengan code unik
                        print(f"Auto-creating dusun: {data['dusun']}")
                        import time
                        import random
                        unique_code = f"D{random.randint(1000, 9999)}"
                        # Pastikan code unik
                        while Dusun.objects.filter(code=unique_code).exists():
                            unique_code = f"D{random.randint(1000, 9999)}"
                        
                        try:
                            dusun = Dusun.objects.create(
                                name=data['dusun'],
                                code=unique_code,
                                description=f"Dusun {data['dusun']} - Dibuat otomatis dari import",
                                area_size=0.0,
                                is_active=True
                            )
                            logger.info(f"Auto-created dusun: {dusun.name}")
                            # Mark as auto-created for tracking
                            dusun._auto_created = True
                            print(f"Successfully created dusun: {dusun.name}")
                        except Exception as e:
                            print(f"ERROR creating dusun: {str(e)}")
                            raise
                    cleaned_data['dusun'] = dusun
            
            # Auto-create lorong jika tidak ada
            lorong = None
            if 'lorong' in data and data['lorong'] and dusun:
                # Import Lorong from references.models
                try:
                    from references.models import Lorong
                except ImportError:
                    Lorong = None
                
                if Lorong is None:
                    print("WARNING: Model Lorong tidak tersedia, skip lorong")
                else:
                    try:
                        lorong = Lorong.objects.get(dusun=dusun, nama_lorong__iexact=data['lorong'])
                    except Lorong.DoesNotExist:
                        # Auto-create lorong baru dengan code unik
                        import random
                        unique_code = f"L{random.randint(1000, 9999)}"
                        # Pastikan code unik
                        while Lorong.objects.filter(kode=unique_code).exists():
                            unique_code = f"L{random.randint(1000, 9999)}"
                            
                        lorong = Lorong.objects.create(
                            nama_lorong=data['lorong'],
                            kode=unique_code,
                            dusun=dusun,
                            rt_number=data.get('rt', '001'),
                            description=f"Lorong {data['lorong']} - Dibuat otomatis dari import",
                            is_active=True
                        )
                        logger.info(f"Auto-created lorong: {lorong.nama_lorong} di dusun {dusun.name}")
                        # Mark as auto-created for tracking
                        lorong._auto_created = True
                    cleaned_data['lorong'] = lorong
            
            # Auto-create RW jika tidak ada
            rw = None
            if 'rw' in data and data['rw'] and dusun:
                # Import RW from references.models
                try:
                    from references.models import RW
                except ImportError:
                    RW = None
                
                if RW is None:
                    print("WARNING: Model RW tidak tersedia, skip rw")
                else:
                    try:
                        rw = RW.objects.get(dusun=dusun, rw_number=str(data['rw']).zfill(3))
                    except RW.DoesNotExist:
                        # Auto-create RW baru
                        rw = RW.objects.create(
                            rw_number=str(data['rw']).zfill(3),
                            dusun=dusun,
                            description=f"RW {data['rw']} - Dibuat otomatis dari import",
                            is_active=True
                        )
                        logger.info(f"Auto-created RW: {rw.rw_number} di dusun {dusun.name}")
                        # Mark as auto-created for tracking
                        rw._auto_created = True
                    cleaned_data['rw'] = rw
            
            # Auto-create RT jika tidak ada
            if 'rt' in data and data['rt'] and rw:
                # Import RT from references.models
                try:
                    from references.models import RT
                except ImportError:
                    RT = None
                
                if RT is None:
                    print("WARNING: Model RT tidak tersedia, skip rt")
                else:
                    try:
                        rt = RT.objects.get(rw=rw, rt_number=str(data['rt']).zfill(3))
                    except RT.DoesNotExist:
                        # Auto-create RT baru
                        rt = RT.objects.create(
                            rt_number=str(data['rt']).zfill(3),
                            rw=rw,
                            description=f"RT {data['rt']} - Dibuat otomatis dari import",
                            is_active=True
                        )
                        logger.info(f"Auto-created RT: {rt.rt_number} di RW {rw.rw_number}")
                        # Mark as auto-created for tracking
                        rt._auto_created = True
                    cleaned_data['rt'] = rt
        
        # Handle RT model specifically
        elif model_class.__name__ == 'RT':
            # Handle RW field for RT model
            rw = None
            if 'rw' in data and data['rw'] and not pd.isna(data['rw']):
                try:
                    from references.models import RW
                    # Try to find RW by name or create if not exists
                    if isinstance(data['rw'], str):
                        # Extract RW number from string like "RW 01 - Dusun Utama"
                        import re
                        rw_match = re.search(r'RW\s+(\d+)', data['rw'])
                        if rw_match:
                            rw_number = rw_match.group(1)
                            try:
                                rw = RW.objects.get(rw_number=rw_number.zfill(3))
                            except RW.DoesNotExist:
                                # Create RW if not exists
                                rw = RW.objects.create(
                                    rw_number=rw_number.zfill(3),
                                    description=f"RW {rw_number} - Dibuat otomatis dari import",
                                    is_active=True
                                )
                                logger.info(f"Auto-created RW: {rw.rw_number}")
                    else:
                        # Try to find by ID
                        try:
                            rw = RW.objects.get(id=int(data['rw']))
                        except (ValueError, RW.DoesNotExist):
                            pass
                except ImportError:
                    print("WARNING: Model RW tidak tersedia")
                
                if rw:
                    cleaned_data['rw'] = rw
                else:
                    print(f"WARNING: RW tidak ditemukan untuk data: {data['rw']}")
                    return None  # Skip this row if RW is required but not found
        
        # Handle RW model specifically
        elif model_class.__name__ == 'RW':
            # Handle Dusun field for RW model
            dusun = None
            if 'dusun' in data and data['dusun'] and not pd.isna(data['dusun']):
                try:
                    from references.models import Dusun
                    # Try to find dusun by name or create if not exists
                    if isinstance(data['dusun'], str):
                        try:
                            dusun = Dusun.objects.get(name__iexact=data['dusun'])
                        except Dusun.DoesNotExist:
                            # Create dusun if not exists
                            import random
                            unique_code = f"D{random.randint(100, 999)}"
                            while Dusun.objects.filter(code=unique_code).exists():
                                unique_code = f"D{random.randint(100, 999)}"
                            
                            dusun = Dusun.objects.create(
                                name=data['dusun'],
                                code=unique_code,
                                description=f"Dusun {data['dusun']} - Dibuat otomatis dari import",
                                is_active=True
                            )
                            logger.info(f"Auto-created dusun: {dusun.name}")
                    else:
                        # Try to find by ID
                        try:
                            dusun = Dusun.objects.get(id=int(data['dusun']))
                        except (ValueError, Dusun.DoesNotExist):
                            pass
                except ImportError:
                    print("WARNING: Model Dusun tidak tersedia")
                
                if dusun:
                    cleaned_data['dusun'] = dusun
                else:
                    print(f"WARNING: Dusun tidak ditemukan untuk data: {data['dusun']}")
                    return None  # Skip this row if Dusun is required but not found
        
        # Handle Lorong model specifically
        elif model_class.__name__ == 'Lorong':
            # Handle Dusun field for Lorong model
            dusun = None
            if 'dusun' in data and data['dusun'] and not pd.isna(data['dusun']):
                try:
                    from references.models import Dusun
                    # Try to find dusun by name or create if not exists
                    if isinstance(data['dusun'], str):
                        try:
                            dusun = Dusun.objects.get(name__iexact=data['dusun'])
                        except Dusun.DoesNotExist:
                            # Create dusun if not exists
                            import random
                            unique_code = f"D{random.randint(100, 999)}"
                            while Dusun.objects.filter(code=unique_code).exists():
                                unique_code = f"D{random.randint(100, 999)}"
                            
                            dusun = Dusun.objects.create(
                                name=data['dusun'],
                                code=unique_code,
                                description=f"Dusun {data['dusun']} - Dibuat otomatis dari import",
                                is_active=True
                            )
                            logger.info(f"Auto-created dusun: {dusun.name}")
                    else:
                        # Try to find by ID
                        try:
                            dusun = Dusun.objects.get(id=int(data['dusun']))
                        except (ValueError, Dusun.DoesNotExist):
                            pass
                except ImportError:
                    print("WARNING: Model Dusun tidak tersedia")
                
                if dusun:
                    cleaned_data['dusun'] = dusun
                else:
                    print(f"WARNING: Dusun tidak ditemukan untuk data: {data['dusun']}")
                    return None  # Skip this row if Dusun is required but not found
        
        # Handle field choices mapping for Penduduk model
        if model_class.__name__ == 'Penduduk':
            # Map education field values
            if 'education' in data and data['education'] and not pd.isna(data['education']):
                education_value = str(data['education']).strip().upper()
                education_mapping = {
                    'TIDAK SEKOLAH': 'TIDAK_BELUM_SEKOLAH',
                    'TIDAK_BELUM_SEKOLAH': 'TIDAK_BELUM_SEKOLAH',
                    'BELUM_TAMAT_SD': 'BELUM_TAMAT_SD',
                    'BELUMTAMATSD': 'BELUM_TAMAT_SD',
                    'SD': 'TAMAT_SD',
                    'TAMATSD': 'TAMAT_SD',
                    'TAMAT_SD': 'TAMAT_SD',
                    'SMP': 'SLTP',
                    'SLTP': 'SLTP',
                    'SMA': 'SLTA',
                    'SLTA': 'SLTA',
                    'D1': 'D1',
                    'D2': 'D2',
                    'D3': 'D3',
                    'D4': 'D4_S1',
                    'D4S1': 'D4_S1',
                    'S1': 'D4_S1',
                    'D4_S1': 'D4_S1',
                    'S2': 'S2',
                    'S3': 'S3'
                }
                data['education'] = education_mapping.get(education_value, 'TIDAK_BELUM_SEKOLAH')
            
            # Map marital_status field values
            if 'marital_status' in data and data['marital_status'] and not pd.isna(data['marital_status']):
                status_value = str(data['marital_status']).strip().upper()
                status_mapping = {
                    'BELUM KAWIN': 'BELUM_KAWIN',
                    'BELUMKAWIN': 'BELUM_KAWIN',
                    'BELUM_KAWIN': 'BELUM_KAWIN',
                    'KAWIN': 'KAWIN',
                    'CERAI HIDUP': 'CERAI_HIDUP',
                    'CERAIHIDUP': 'CERAI_HIDUP',
                    'CERAI_HIDUP': 'CERAI_HIDUP',
                    'CERAI MATI': 'CERAI_MATI',
                    'CERAIMATI': 'CERAI_MATI',
                    'CERAI_MATI': 'CERAI_MATI'
                }
                data['marital_status'] = status_mapping.get(status_value, 'BELUM_KAWIN')
        
        for field_name, value in data.items():
            if field_name in model_fields:
                field = model_fields[field_name]
                
                # Skip auto-generated timestamp fields
                if field_name in ['created_at', 'updated_at'] and field.auto_now:
                    continue
                
                # Handle different field types
                if field.__class__.__name__ == 'ForeignKey':
                    # Handle foreign key fields
                    if field_name == 'dusun_id' and value:
                        try:
                            # Import Dusun safely
                            try:
                                from references.models import Dusun
                            except ImportError:
                                Dusun = None
                            
                            if Dusun is not None:
                                dusun = Dusun.objects.get(id=int(value))
                                cleaned_data['dusun'] = dusun
                        except:
                            pass
                    elif field_name == 'penduduk_id' and value:
                        try:
                            penduduk = Penduduk.objects.get(id=int(value))
                            cleaned_data['penduduk'] = penduduk
                        except:
                            pass
                    elif field_name == 'rw_id' and value:
                        try:
                            # Import RW safely
                            try:
                                from references.models import RW
                            except ImportError:
                                RW = None
                            
                            if RW is not None:
                                rw = RW.objects.get(id=int(value))
                                cleaned_data['rw'] = rw
                        except:
                            pass
                elif field.__class__.__name__ == 'DateField' and value:
                    # Handle date fields
                    if pd.isna(value):
                        cleaned_data[field_name] = None
                    elif isinstance(value, str):
                        try:
                            cleaned_data[field_name] = pd.to_datetime(value).date()
                        except:
                            cleaned_data[field_name] = None
                    elif hasattr(value, 'date'):
                        cleaned_data[field_name] = value.date()
                    else:
                        cleaned_data[field_name] = value
                elif field.__class__.__name__ == 'BooleanField' and value:
                    # Handle boolean fields
                    if isinstance(value, str):
                        cleaned_data[field_name] = value.lower() in ['ya', 'yes', 'true', '1']
                    else:
                        cleaned_data[field_name] = bool(value)
                else:
                    # Handle NaN and null values
                    if pd.isna(value) or value is None or value == '':
                        if field.null:
                            cleaned_data[field_name] = None
                        else:
                            # Skip this field if it's required but empty
                            continue
                    else:
                        # Convert value to appropriate type
                        if field.__class__.__name__ == 'CharField':
                            cleaned_data[field_name] = str(value).strip()
                        elif field.__class__.__name__ == 'IntegerField':
                            try:
                                cleaned_data[field_name] = int(value)
                            except (ValueError, TypeError):
                                if field.null:
                                    cleaned_data[field_name] = None
                                else:
                                    continue
                        elif field.__class__.__name__ == 'DecimalField':
                            try:
                                cleaned_data[field_name] = float(value)
                            except (ValueError, TypeError):
                                if field.null:
                                    cleaned_data[field_name] = None
                                else:
                                    continue
                        else:
                            cleaned_data[field_name] = value
        
        return cleaned_data
    
    def _export_to_pdf(self, model_name: str, queryset, include_related: bool = True) -> HttpResponse:
        """Export data ke PDF dengan styling yang lebih baik"""
        if not PDF_AVAILABLE:
            raise ValueError("PDF export tidak tersedia. Install reportlab: pip install reportlab")
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            rightMargin=50, 
            leftMargin=50, 
            topMargin=50, 
            bottomMargin=50
        )
        
        # Untuk penduduk, hanya tampilkan 4 kolom penting saja
        if model_name == 'penduduk':
            # Hanya 4 kolom penting untuk penduduk
            data = []
            for obj in queryset:  # Ambil semua data penduduk
                row = [
                    str(obj.nik),  # Pastikan NIK sebagai string
                    obj.name[:30] + '...' if len(obj.name) > 30 else obj.name,
                    obj.get_gender_display(),
                    obj.dusun.name if obj.dusun else '-'
                ]
                data.append(row)
            
            # Create table with simplified structure - 4 kolom saja
            table_data = [['NIK', 'Nama', 'Jenis Kelamin', 'Dusun']]
            
            for row in data:
                table_data.append(row)
            
            # Column widths untuk penduduk - 4 kolom saja
            col_widths = [1.8 * inch, 2.5 * inch, 1.2 * inch, 2.0 * inch]
        else:
            # Untuk model lain, gunakan konfigurasi default
            field_config = self._get_field_config(model_name)
            
            # Prepare data
            data = []
            for obj in queryset:  # Ambil semua data
                row = []
                for config in field_config[:5]:  # Hanya 5 kolom pertama
                    value = self._get_field_value(obj, config['field_name'], include_related)
                    # Convert datetime objects to string
                    if isinstance(value, (datetime, date)):
                        value = value.strftime('%Y-%m-%d')
                    # Truncate long text
                    if isinstance(value, str) and len(value) > 25:
                        value = value[:22] + '...'
                    row.append(str(value) if value is not None else '-')
                data.append(row)
            
            # Create table with better structure
            table_data = [['No'] + [config['verbose_name'] for config in field_config[:5]]]
            
            for i, row in enumerate(data, 1):
                table_data.append([str(i)] + row)
            
            # Create table with better column widths
            col_widths = [0.5 * inch]  # No column
            for config in field_config[:5]:
                col_widths.append(1.2 * inch)
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Enhanced table styling
        table_style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),  # Professional blue
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),  # Light gray
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2E86AB')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ])
        
        table.setStyle(table_style)
        
        # Build PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Header with logo and title
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2E86AB'),
            fontName='Helvetica-Bold'
        )
        
        # Subtitle style
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        
        # Info style
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=2,  # Right alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        
        # Add header
        elements.append(Paragraph("LAPORAN DATA PENDUDUK", header_style))
        elements.append(Paragraph("Desa Pulosarok", subtitle_style))
        elements.append(Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y, %H:%M WIB')}", info_style))
        elements.append(Spacer(1, 20))
        
        # Add table
        elements.append(table)
        
        # Add footer info
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        elements.append(Paragraph(f"Total Data: {len(data)} | Halaman 1", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF content
        buffer.seek(0)
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Laporan_Data_{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
    
    def _detect_and_map_penduduk_columns(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Deteksi dan mapping kolom otomatis untuk data penduduk"""
        if not data:
            return data
        
        # Ambil sample data untuk deteksi
        sample_row = data[0]
        original_columns = list(sample_row.keys())
        
        # Mapping kolom yang mungkin
        column_mappings = {
            # NIK variations
            'nik': ['nik', 'NIK', 'no_ktp', 'ktp', 'no_identitas', 'identitas'],
            # Nama variations  
            'name': ['name', 'nama', 'Nama', 'nama_lengkap', 'full_name', 'nama_penduduk', 'Nama Lengkap'],
            # Gender variations
            'gender': ['gender', 'jenis_kelamin', 'jk', 'sex', 'kelamin', 'Jenis Kelamin'],
            # Tempat lahir variations
            'birth_place': ['birth_place', 'tempat_lahir', 'tempat_lahir', 'place_of_birth', 'Tempat Lahir'],
            # Tanggal lahir variations
            'birth_date': ['birth_date', 'tanggal_lahir', 'tgl_lahir', 'date_of_birth', 'lahir', 'Tanggal Lahir'],
            # KK variations
            'kk_number': ['kk_number', 'no_kk', 'kk', 'nomor_kk', 'no_kartu_keluarga', 'Nomor KK'],
            # Agama variations
            'religion': ['religion', 'agama', 'religi', 'Agama'],
            # Pendidikan variations
            'education': ['education', 'pendidikan', 'pendidikan_terakhir', 'pendidikan_tertinggi', 'Pendidikan'],
            # Pekerjaan variations
            'occupation': ['occupation', 'pekerjaan', 'job', 'profesi', 'Pekerjaan'],
            # Status perkawinan variations
            'marital_status': ['marital_status', 'status_perkawinan', 'status_kawin', 'kawin', 'Status Perkawinan'],
            # Alamat variations
            'address': ['address', 'alamat', 'alamat_lengkap', 'full_address', 'Alamat'],
            # Telepon variations
            'phone_number': ['phone_number', 'telepon', 'telp', 'no_telp', 'phone', 'hp', 'Telepon'],
            # RT variations
            'rt': ['rt', 'RT', 'rt_number', 'no_rt', 'nomor_rt'],
            # RW variations
            'rw': ['rw', 'RW', 'rw_number', 'no_rw', 'nomor_rw'],
            # Dusun variations
            'dusun': ['dusun', 'Dusun', 'nama_dusun', 'dusun_name', 'Dusun'],
            # Lorong variations
            'lorong': ['lorong', 'Lorong', 'nama_lorong', 'lorong_name', 'Lorong'],
            # Hubungan kepala keluarga variations
            'relationship_to_head': ['relationship_to_head', 'hubungan_kepala_keluarga', 'hubungan_kk', 'status_keluarga', 'Hubungan dengan Kepala Keluarga'],
        }
        
        # Buat mapping dari kolom asli ke kolom standar
        column_map = {}
        for standard_field, possible_columns in column_mappings.items():
            for original_col in original_columns:
                if original_col.lower() in [col.lower() for col in possible_columns]:
                    column_map[original_col] = standard_field
                    break
        
        # Transform data dengan mapping
        transformed_data = []
        for row in data:
            new_row = {}
            for original_col, value in row.items():
                if original_col in column_map:
                    new_row[column_map[original_col]] = value
                else:
                    # Jika tidak ada mapping, gunakan nama asli
                    new_row[original_col] = value
            transformed_data.append(new_row)
        
        logger.info(f"Column mapping applied: {column_map}")
        return transformed_data

