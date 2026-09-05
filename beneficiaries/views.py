from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.http import JsonResponse
from django.views.generic import ListView, DetailView
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from .models import (
    Beneficiary, BeneficiaryCategory, Aid, AidDistribution,
    BeneficiaryVerification, TarafKehidupan, DataBantuan,
    DokumenGampong, Berita
)
# Use references.models.Penduduk
from references.models import Penduduk


def beneficiaries_index(request):
    """Halaman utama data penerima bantuan"""
    # Statistik umum
    total_beneficiaries = Beneficiary.objects.filter(status='aktif').count()
    total_programs = Aid.objects.filter(is_active=True).count()
    total_distributed = AidDistribution.objects.filter(status='distributed').count()
    
    context = {
        'total_beneficiaries': total_beneficiaries,
        'total_programs': total_programs,
        'total_distributed': total_distributed,
    }
    
    return render(request, 'public/beneficiaries/index.html', context)


def beneficiaries_list(request):
    """Daftar penerima bantuan untuk publik"""
    beneficiaries = Beneficiary.objects.filter(
        status='aktif'
    ).select_related('person', 'category').order_by('-created_at')
    
    # Filter berdasarkan kategori
    category_id = request.GET.get('category')
    if category_id:
        beneficiaries = beneficiaries.filter(category_id=category_id)
    
    # Filter berdasarkan status ekonomi
    economic_status = request.GET.get('economic_status')
    if economic_status:
        beneficiaries = beneficiaries.filter(economic_status=economic_status)
    
    # Search
    search = request.GET.get('search')
    if search:
        beneficiaries = beneficiaries.filter(
            Q(person__name__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(beneficiaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Data untuk filter
    categories = BeneficiaryCategory.objects.filter(is_active=True)
    economic_statuses = Beneficiary.ECONOMIC_STATUS_CHOICES
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'economic_statuses': economic_statuses,
        'current_category': category_id,
        'current_economic_status': economic_status,
        'search_query': search,
        'total_beneficiaries': beneficiaries.count()
    }
    
    return render(request, 'public/beneficiaries/beneficiaries_list.html', context)


def beneficiary_categories(request):
    """Daftar kategori penerima bantuan"""
    categories = BeneficiaryCategory.objects.filter(
        is_active=True
    ).annotate(
        beneficiary_count=Count('beneficiary')
    ).order_by('name')
    
    context = {
        'categories': categories
    }
    
    return render(request, 'public/beneficiary_categories.html', context)


def category_detail(request, category_id):
    """Detail kategori penerima bantuan"""
    category = get_object_or_404(BeneficiaryCategory, id=category_id, is_active=True)
    
    beneficiaries = Beneficiary.objects.filter(
        category=category,
        status='aktif'
    ).select_related('person').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(beneficiaries, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistik
    stats = {
        'total_beneficiaries': beneficiaries.count(),
        'by_economic_status': beneficiaries.values('economic_status').annotate(
            count=Count('id')
        ).order_by('economic_status')
    }
    
    context = {
        'category': category,
        'page_obj': page_obj,
        'stats': stats
    }
    
    return render(request, 'public/category_detail.html', context)


def aid_programs(request):
    """Daftar program bantuan"""
    aids = Aid.objects.filter(
        is_active=True,
        start_date__lte=timezone.now().date()
    ).order_by('-start_date')
    
    # Filter berdasarkan jenis bantuan
    aid_type = request.GET.get('aid_type')
    if aid_type:
        aids = aids.filter(aid_type=aid_type)
    
    # Filter berdasarkan sumber
    source = request.GET.get('source')
    if source:
        aids = aids.filter(source=source)
    
    # Search
    search = request.GET.get('search')
    if search:
        aids = aids.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(aids, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Data untuk filter
    aid_types = Aid.AID_TYPE_CHOICES
    sources = Aid.SOURCE_CHOICES
    
    context = {
        'page_obj': page_obj,
        'aid_types': aid_types,
        'sources': sources,
        'current_aid_type': aid_type,
        'current_source': source,
        'search_query': search
    }
    
    return render(request, 'public/aid_list.html', context)


def aid_detail(request, aid_id):
    """Detail program bantuan"""
    aid = get_object_or_404(Aid, id=aid_id, is_active=True)
    
    # Statistik distribusi
    distributions = AidDistribution.objects.filter(aid=aid)
    stats = {
        'total_distributed': distributions.filter(status='distributed').count(),
        'total_approved': distributions.filter(status='approved').count(),
        'total_pending': distributions.filter(status='pending').count(),
        'total_amount_distributed': distributions.filter(
            status='distributed'
        ).aggregate(total=Sum('amount_received'))['total'] or 0,
        'progress_percentage': (
            distributions.filter(status='distributed').count() / 
            aid.target_beneficiaries * 100
        ) if aid.target_beneficiaries > 0 else 0
    }
    
    # Daftar penerima (hanya yang sudah disalurkan)
    recent_distributions = distributions.filter(
        status='distributed'
    ).select_related('beneficiary__person').order_by('-distribution_date')[:10]
    
    context = {
        'aid': aid,
        'stats': stats,
        'recent_distributions': recent_distributions
    }
    
    return render(request, 'public/aid_detail.html', context)


def aid_statistics(request):
    """Statistik bantuan untuk publik"""
    # Statistik umum
    total_aids = Aid.objects.filter(is_active=True).count()
    total_beneficiaries = Beneficiary.objects.filter(status='aktif').count()
    total_distributed = AidDistribution.objects.filter(status='distributed').count()
    total_budget = Aid.objects.filter(is_active=True).aggregate(
        total=Sum('total_budget')
    )['total'] or 0
    
    # Statistik per jenis bantuan
    aid_by_type = Aid.objects.filter(is_active=True).values(
        'aid_type'
    ).annotate(
        count=Count('id'),
        total_budget=Sum('total_budget')
    ).order_by('-count')
    
    # Statistik per sumber
    aid_by_source = Aid.objects.filter(is_active=True).values(
        'source'
    ).annotate(
        count=Count('id'),
        total_budget=Sum('total_budget')
    ).order_by('-count')
    
    # Statistik per kategori penerima
    beneficiaries_by_category = BeneficiaryCategory.objects.filter(
        is_active=True
    ).annotate(
        beneficiary_count=Count('beneficiary')
    ).order_by('-beneficiary_count')
    
    context = {
        'total_aids': total_aids,
        'total_beneficiaries': total_beneficiaries,
        'total_distributed': total_distributed,
        'total_budget': total_budget,
        'aid_by_type': aid_by_type,
        'aid_by_source': aid_by_source,
        'beneficiaries_by_category': beneficiaries_by_category
    }
    
    return render(request, 'public/aid_statistics.html', context)


def data_bantuan_list(request):
    """Daftar data bantuan per penduduk"""
    data_bantuan = DataBantuan.objects.filter(
        status='aktif'
    ).select_related('person').order_by('-tanggal_mulai')
    
    # Filter berdasarkan jenis bantuan
    jenis_bantuan = request.GET.get('jenis_bantuan')
    if jenis_bantuan:
        data_bantuan = data_bantuan.filter(jenis_bantuan=jenis_bantuan)
    
    # Search
    search = request.GET.get('search')
    if search:
        data_bantuan = data_bantuan.filter(
            Q(person__name__icontains=search) |
            Q(nama_program__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(data_bantuan, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Data untuk filter
    jenis_bantuan_choices = DataBantuan.JENIS_BANTUAN_CHOICES
    
    context = {
        'page_obj': page_obj,
        'jenis_bantuan_choices': jenis_bantuan_choices,
        'current_jenis_bantuan': jenis_bantuan,
        'search_query': search
    }
    
    return render(request, 'public/data_bantuan_list.html', context)


def documents_list(request):
    """Daftar dokumen publik"""
    documents = DokumenGampong.objects.filter(
        is_public=True,
        status='approved'
    ).order_by('-created_at')
    
    # Filter berdasarkan kategori
    kategori = request.GET.get('kategori')
    if kategori:
        documents = documents.filter(kategori=kategori)
    
    # Search
    search = request.GET.get('search')
    if search:
        documents = documents.filter(
            Q(nama_dokumen__icontains=search) |
            Q(deskripsi__icontains=search) |
            Q(tags__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(documents, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Data untuk filter
    kategori_choices = DokumenGampong.KATEGORI_CHOICES
    
    context = {
        'page_obj': page_obj,
        'kategori_choices': kategori_choices,
        'current_kategori': kategori,
        'search_query': search
    }
    
    return render(request, 'public/documents_list.html', context)


def news_list(request):
    """Daftar berita"""
    news = Berita.objects.filter(
        status='published'
    ).order_by('-published_at')
    
    # Filter berdasarkan kategori
    kategori = request.GET.get('kategori')
    if kategori:
        news = news.filter(kategori=kategori)
    
    # Search
    search = request.GET.get('search')
    if search:
        news = news.filter(
            Q(title__icontains=search) |
            Q(content__icontains=search) |
            Q(excerpt__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(news, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Berita unggulan
    featured_news = Berita.objects.filter(
        status='published',
        is_featured=True
    ).order_by('-published_at')[:3]
    
    # Data untuk filter
    kategori_choices = Berita.KATEGORI_CHOICES
    
    context = {
        'page_obj': page_obj,
        'featured_news': featured_news,
        'kategori_choices': kategori_choices,
        'current_kategori': kategori,
        'search_query': search
    }
    
    return render(request, 'public/news_list.html', context)


def news_detail(request, slug):
    """Detail berita"""
    news = get_object_or_404(Berita, slug=slug, status='published')
    
    # Increment views count
    news.views_count += 1
    news.save(update_fields=['views_count'])
    
    # Berita terkait
    related_news = Berita.objects.filter(
        status='published',
        kategori=news.kategori
    ).exclude(id=news.id).order_by('-published_at')[:4]
    
    context = {
        'news': news,
        'related_news': related_news
    }
    
    return render(request, 'public/news_detail.html', context)


# API Views untuk AJAX
def api_beneficiaries_list(request):
    """API untuk daftar penerima bantuan - optimized for mobile"""
    # Get page size based on device (smaller for mobile)
    page_size = int(request.GET.get('page_size', 12))
    page_size = min(page_size, 20)  # Limit max page size for performance
    
    beneficiaries = Beneficiary.objects.filter(
        status='aktif'
    ).select_related('person', 'category').only(
        'id', 'economic_status', 'status', 'notes', 'created_at',
        'person__nama', 'person__nik', 'person__alamat',
        'category__id', 'category__name'
    ).order_by('-created_at')
    
    # Filter berdasarkan kategori
    category_id = request.GET.get('category')
    if category_id:
        beneficiaries = beneficiaries.filter(category_id=category_id)
    
    # Filter berdasarkan status ekonomi
    economic_status = request.GET.get('economic_status')
    if economic_status:
        beneficiaries = beneficiaries.filter(economic_status=economic_status)
    
    # Search with optimized query
    search = request.GET.get('search')
    if search:
        beneficiaries = beneficiaries.filter(
            Q(person__name__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(beneficiaries, page_size)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Serialize data with minimal fields for mobile
    results = []
    for beneficiary in page_obj:
        results.append({
            'id': beneficiary.id,
            'person': {
                'name': beneficiary.person.name,
                'nik': beneficiary.person.nik,
                'address': beneficiary.person.address,
            },
            'category': {
                'id': beneficiary.category.id,
                'name': beneficiary.category.name,
            },
            'economic_status': beneficiary.economic_status,
            'economic_status_display': beneficiary.get_economic_status_display(),
            'status': beneficiary.status,
            'status_display': beneficiary.get_status_display(),
            'notes': beneficiary.notes[:100] + '...' if beneficiary.notes and len(beneficiary.notes) > 100 else beneficiary.notes,
        })
    
    return JsonResponse({
        'results': results,
        'count': paginator.count,
        'num_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        'page_size': page_size,
    })


def api_beneficiary_detail(request, beneficiary_id):
    """API untuk detail penerima bantuan"""
    beneficiary = get_object_or_404(Beneficiary, id=beneficiary_id, status='aktif')
    
    data = {
        'id': beneficiary.id,
        'person': {
            'name': beneficiary.person.name,
            'nik': beneficiary.person.nik,
            'address': beneficiary.person.address,
            'birth_place': beneficiary.person.birth_place,
            'birth_date': beneficiary.person.birth_date.strftime('%d-%m-%Y') if beneficiary.person.birth_date else None,
            'gender': beneficiary.person.gender,
        },
        'category': {
            'id': beneficiary.category.id,
            'name': beneficiary.category.name,
            'description': beneficiary.category.description,
        },
        'economic_status': beneficiary.economic_status,
        'economic_status_display': beneficiary.get_economic_status_display(),
        'status': beneficiary.status,
        'status_display': beneficiary.get_status_display(),
        'notes': beneficiary.notes,
        'created_at': beneficiary.created_at.strftime('%d-%m-%Y %H:%M'),
    }
    
    return JsonResponse(data)


def api_categories_list(request):
    """API untuk daftar kategori bantuan"""
    categories = BeneficiaryCategory.objects.filter(
        is_active=True
    ).annotate(
        beneficiary_count=Count('beneficiary')
    ).order_by('name')
    
    results = []
    for category in categories:
        results.append({
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'beneficiary_count': category.beneficiary_count,
        })
    
    return JsonResponse(results, safe=False)


def api_beneficiaries_stats(request):
    """API untuk statistik penerima bantuan"""
    stats = {
        'total_beneficiaries': Beneficiary.objects.filter(status='aktif').count(),
        'by_category': list(BeneficiaryCategory.objects.filter(
            is_active=True
        ).annotate(
            count=Count('beneficiary')
        ).values('name', 'count')),
        'by_economic_status': list(Beneficiary.objects.filter(
            status='aktif'
        ).values('economic_status').annotate(
            count=Count('id')
        ).values('economic_status', 'count'))
    }
    
    return JsonResponse(stats)

# ==================== ADMIN PANEL INTEGRATION VIEWS ====================

from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.serializers import serialize
from django.forms.models import model_to_dict
import json
from datetime import datetime, timedelta

def is_admin(user):
    """Check if user is admin"""
    return user.is_staff or user.is_superuser

@login_required
@user_passes_test(is_admin)
def admin_beneficiaries_dashboard(request):
    """Admin dashboard for beneficiaries management"""
    # Statistics for dashboard
    total_beneficiaries = Beneficiary.objects.count()
    active_beneficiaries = Beneficiary.objects.filter(status='aktif').count()
    pending_verifications = BeneficiaryVerification.objects.filter(verification_status='pending').count()
    total_aid_programs = Aid.objects.count()
    active_aid_programs = Aid.objects.filter(is_active=True).count()
    total_distributions = AidDistribution.objects.count()
    pending_distributions = AidDistribution.objects.filter(status='pending').count()
    
    # Recent activities
    recent_beneficiaries = Beneficiary.objects.select_related('person', 'category').order_by('-created_at')[:10]
    recent_distributions = AidDistribution.objects.select_related('beneficiary', 'aid').order_by('-created_at')[:10]
    recent_verifications = BeneficiaryVerification.objects.select_related('beneficiary').order_by('-created_at')[:10]
    
    # Monthly statistics
    current_month = timezone.now().replace(day=1)
    monthly_registrations = Beneficiary.objects.filter(
        created_at__gte=current_month
    ).count()
    monthly_distributions = AidDistribution.objects.filter(
        created_at__gte=current_month,
        status='distributed'
    ).count()
    
    # Category statistics (limit to 10)
    category_stats = BeneficiaryCategory.objects.annotate(
        total_beneficiaries=Count('beneficiary'),
        active_beneficiaries=Count('beneficiary', filter=Q(beneficiary__status='aktif')),
        inactive_beneficiaries=Count('beneficiary', filter=Q(beneficiary__status='nonaktif'))
    ).order_by('-total_beneficiaries')[:10]
    
    # Inactive beneficiaries count
    inactive_beneficiaries = Beneficiary.objects.filter(status='nonaktif').count()
    
    # Monthly distribution data for chart (last 6 months)
    from datetime import timedelta
    from django.db.models.functions import TruncMonth
    
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_data = AidDistribution.objects.filter(
        created_at__gte=six_months_ago,
        status='distributed'
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Format monthly data for chart
    month_labels = []
    month_values = []
    for entry in monthly_data:
        month_labels.append(entry['month'].strftime('%b'))
        month_values.append(entry['count'])
    
    # Fill missing months with 0
    if len(month_labels) < 6:
        default_months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun']
        default_values = [0] * 6
        # Use actual data if available
        if month_labels:
            month_labels = default_months
            month_values = default_values
    
    # Total categories
    total_categories = BeneficiaryCategory.objects.count()
    
    context = {
        'title': 'Dashboard Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'beneficiaries_dashboard',
        'total_beneficiaries': total_beneficiaries,
        'active_beneficiaries': active_beneficiaries,
        'inactive_beneficiaries': inactive_beneficiaries,
        'pending_verifications': pending_verifications,
        'total_categories': total_categories,
        'total_aid_programs': total_aid_programs,
        'active_aid_programs': active_aid_programs,
        'total_distributions': total_distributions,
        'pending_distributions': pending_distributions,
        'recent_beneficiaries': recent_beneficiaries,
        'recent_distributions': recent_distributions,
        'recent_verifications': recent_verifications,
        'monthly_registrations': monthly_registrations,
        'monthly_distributions': monthly_distributions,
        'category_stats': category_stats,
        'month_labels': month_labels,
        'month_values': month_values,
    }
    return render(request, 'admin_panel/beneficiaries/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def admin_beneficiaries_list(request):
    """Admin list view for beneficiaries with advanced filtering"""
    
    # Handle bulk delete POST request
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'bulk_delete':
            try:
                ids = request.POST.getlist('ids')
                if not ids:
                    return JsonResponse({
                        'success': False,
                        'message': 'Tidak ada data yang dipilih'
                    })
                
                # Convert to integers
                ids = [int(id) for id in ids]
                
                # Get beneficiaries to delete
                beneficiaries_to_delete = Beneficiary.objects.filter(id__in=ids)
                deleted_count = beneficiaries_to_delete.count()
                
                # Delete the beneficiaries
                beneficiaries_to_delete.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Berhasil menghapus {deleted_count} penerima bantuan',
                    'deleted_count': deleted_count
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error: {str(e)}'
                })
    
    beneficiaries = Beneficiary.objects.select_related('person', 'category').all()
    
    # Advanced filtering
    status = request.GET.get('status')
    if status:
        beneficiaries = beneficiaries.filter(status=status)
    
    category_id = request.GET.get('category')
    if category_id:
        beneficiaries = beneficiaries.filter(category_id=category_id)
    
    economic_status = request.GET.get('economic_status')
    if economic_status:
        beneficiaries = beneficiaries.filter(economic_status=economic_status)
    
    verification_status = request.GET.get('verification_status')
    if verification_status:
        beneficiaries = beneficiaries.filter(verification_status=verification_status)
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        beneficiaries = beneficiaries.filter(created_at__date__gte=date_from)
    if date_to:
        beneficiaries = beneficiaries.filter(created_at__date__lte=date_to)
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        beneficiaries = beneficiaries.filter(
            Q(person__name__icontains=search) |
            Q(person__nik__icontains=search) |
            Q(address__icontains=search) |
            Q(phone__icontains=search)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    if sort_by in ['name', '-name', 'created_at', '-created_at', 'status', '-status']:
        if sort_by in ['name', '-name']:
            sort_by = sort_by.replace('name', 'person__nama')
        beneficiaries = beneficiaries.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(beneficiaries, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get categories for filter dropdown
    categories = BeneficiaryCategory.objects.all()
    
    context = {
        'title': 'Manajemen Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'beneficiaries_list',
        'beneficiaries': page_obj,
        'categories': categories,
        'current_filters': {
            'status': status,
            'category': category_id,
            'economic_status': economic_status,
            'verification_status': verification_status,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'sort': sort_by,
        }
    }
    return render(request, 'admin_panel/beneficiaries/list.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_beneficiary_create(request):
    """Create new beneficiary via admin panel"""
    if request.method == 'GET':
        # Return HTML form for GET request
        from .forms import BeneficiaryForm
        form = BeneficiaryForm()
        categories = BeneficiaryCategory.objects.all()
        context = {
            'active_menu': 'beneficiaries',
            'active_submenu': 'beneficiaries_add',
            'form': form,
            'categories': categories,
            'economic_status_choices': Beneficiary.ECONOMIC_STATUS_CHOICES,
        }
        return render(request, 'admin_panel/beneficiaries/form.html', context)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validate penduduk_id for new beneficiaries
            penduduk_id = data.get('penduduk_id')
            if not penduduk_id or penduduk_id == '':
                return JsonResponse({
                    'success': False,
                    'message': 'Data penduduk harus dipilih'
                })
            
            # Validate category_id
            category_id = data.get('category_id')
            if not category_id or category_id == '':
                return JsonResponse({
                    'success': False,
                    'message': 'Kategori bantuan harus dipilih'
                })
            
            try:
                category_id = int(category_id)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'Kategori bantuan tidak valid'
                })
            
            # Get person record by ID
            try:
                person = Penduduk.objects.get(id=penduduk_id)
            except Penduduk.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Data penduduk tidak ditemukan'
                })
            
            # Check if person is already registered for this category
            existing_beneficiary = Beneficiary.objects.filter(
                person=person, 
                category_id=category_id
            ).first()
            
            if existing_beneficiary:
                return JsonResponse({
                    'success': False,
                    'message': f'{person.name} sudah terdaftar sebagai penerima bantuan untuk kategori "{existing_beneficiary.category.name}". Silakan pilih kategori lain atau edit data yang sudah ada.'
                })
            
            # Create beneficiary record
            beneficiary = Beneficiary.objects.create(
                person=person,
                category_id=category_id,
                registration_date=data.get('registration_date'),
                economic_status=data.get('economic_status', 'miskin'),
                monthly_income=data.get('monthly_income', 0),
                family_members_count=data.get('family_members_count', 1),
                house_condition=data.get('house_condition', ''),
                special_needs=data.get('special_needs', ''),
                notes=data.get('notes', ''),
                status=data.get('status', 'aktif')
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Penerima bantuan berhasil ditambahkan',
                'data': {
                    'id': beneficiary.id,
                    'name': beneficiary.person.name,
                    'nik': beneficiary.person.nik,
                    'category': beneficiary.category.name if beneficiary.category else None,
                    'status': beneficiary.status
                }
            })
        except Exception as e:
            error_message = str(e)
            
            # Handle specific database errors
            if 'UNIQUE constraint failed' in error_message and 'person_id' in error_message:
                return JsonResponse({
                    'success': False,
                    'message': 'Penduduk ini sudah terdaftar sebagai penerima bantuan untuk kategori yang dipilih. Silakan pilih kategori lain atau edit data yang sudah ada.'
                })
            elif 'IntegrityError' in error_message:
                return JsonResponse({
                    'success': False,
                    'message': 'Terjadi kesalahan integritas data. Pastikan data yang dimasukkan valid dan tidak duplikat.'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'Error: {error_message}'
                })
    
    # GET request - return form data
    categories = BeneficiaryCategory.objects.all()
    return JsonResponse({
        'categories': [{'id': c.id, 'name': c.name} for c in categories],
        'economic_status_choices': Beneficiary.ECONOMIC_STATUS_CHOICES,
    })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_beneficiary_update(request, pk):
    """Update beneficiary via admin panel"""
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Update person data if provided
            person_data = data.get('person', {})
            if person_data:
                person = beneficiary.person
                person.name = person_data.get('name', person.name)
                person.gender = person_data.get('gender', person.gender)
                person.birth_date = person_data.get('birth_date', person.birth_date)
                person.birth_place = person_data.get('birth_place', person.birth_place)
                person.religion = person_data.get('religion', person.religion)
                person.education = person_data.get('education', person.education)
                person.occupation = person_data.get('occupation', person.occupation)
                person.marital_status = person_data.get('marital_status', person.marital_status)
                person.save()
            
            # Update beneficiary data
            beneficiary.category_id = data.get('category_id', beneficiary.category_id)
            beneficiary.economic_status = data.get('economic_status', beneficiary.economic_status)
            beneficiary.family_members_count = data.get('family_members_count', beneficiary.family_members_count)
            beneficiary.monthly_income = data.get('monthly_income', beneficiary.monthly_income)
            beneficiary.house_condition = data.get('house_condition', beneficiary.house_condition)
            beneficiary.special_needs = data.get('special_needs', beneficiary.special_needs)
            beneficiary.notes = data.get('notes', beneficiary.notes)
            beneficiary.status = data.get('status', beneficiary.status)
            beneficiary.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Data penerima bantuan berhasil diperbarui',
                'data': {
                    'id': beneficiary.id,
                    'name': beneficiary.person.name,
                    'nik': beneficiary.person.nik,
                    'category': beneficiary.category.name if beneficiary.category else None,
                    'status': beneficiary.status
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    # GET request - return HTML form
    categories = BeneficiaryCategory.objects.all()
    context = {
        'title': 'Edit Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'beneficiaries',
        'beneficiary': beneficiary,
        'categories': categories,
        'economic_status_choices': Beneficiary.ECONOMIC_STATUS_CHOICES,
    }
    return render(request, 'admin_panel/beneficiaries/form.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST", "DELETE"])
def admin_beneficiary_delete(request, pk):
    """Delete beneficiary via admin panel"""
    try:
        beneficiary = get_object_or_404(Beneficiary, pk=pk)
        name = beneficiary.person.name
        beneficiary.delete()
        return JsonResponse({
            'success': True,
            'message': f'Penerima bantuan {name} berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@user_passes_test(is_admin)
def admin_beneficiary_detail(request, pk):
    """Detailed view of beneficiary for admin"""
    beneficiary = get_object_or_404(Beneficiary.objects.select_related('person', 'category'), pk=pk)
    
    # Get related data
    verifications = BeneficiaryVerification.objects.filter(beneficiary=beneficiary).order_by('-created_at')
    distributions = AidDistribution.objects.filter(beneficiary=beneficiary).select_related('aid').order_by('-created_at')
    
    context = {
        'title': f'Detail Penerima Bantuan - {beneficiary.person.name}',
        'active_menu': 'beneficiaries',
        'active_submenu': 'beneficiaries_detail',
        'beneficiary': beneficiary,
        'verifications': verifications,
        'distributions': distributions,
    }
    return render(request, 'admin_panel/beneficiaries/detail.html', context)

# ==================== MISSING VIEW FUNCTIONS ====================

@login_required
@user_passes_test(is_admin)
def beneficiary_detail(request, pk):
    """Detail view for beneficiary"""
    return admin_beneficiary_detail(request, pk)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def beneficiary_update(request, pk):
    """Update beneficiary"""
    return admin_beneficiary_update(request, pk)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST", "DELETE"])
def beneficiary_delete(request, pk):
    """Delete beneficiary"""
    return admin_beneficiary_delete(request, pk)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["DELETE"])
def aid_distribution_delete(request, pk):
    """Delete aid distribution"""
    try:
        from beneficiaries.models import AidDistribution
        distribution = get_object_or_404(AidDistribution, pk=pk)
        recipient_name = distribution.beneficiary.person.name
        distribution.delete()
        return JsonResponse({
            'success': True,
            'message': f'Distribusi bantuan untuk {recipient_name} berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== CATEGORY MANAGEMENT ====================

@login_required
@user_passes_test(is_admin)
def admin_categories_list(request):
    """Admin list view for beneficiary categories"""
    categories = BeneficiaryCategory.objects.annotate(
        beneficiary_count=Count('beneficiary')
    ).order_by('name')
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        categories = categories.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(categories, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Manajemen Kategori Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'categories',
        'categories': page_obj,
        'search': search,
    }
    return render(request, 'admin_panel/beneficiaries/categories.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_category_create(request):
    """Create new beneficiary category"""
    if request.method == 'GET':
        # Return HTML form for GET request
        from .forms import BeneficiaryCategoryForm
        form = BeneficiaryCategoryForm()
        context = {
            'title': 'Tambah Kategori Penerima Bantuan',
            'active_menu': 'beneficiaries',
            'active_submenu': 'categories',
            'form': form,
        }
        return render(request, 'admin_panel/beneficiaries/category_form.html', context)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            category = BeneficiaryCategory.objects.create(
                name=data.get('name'),
                description=data.get('description', ''),
                criteria=data.get('criteria', ''),
                is_active=data.get('is_active', True)
            )
            return JsonResponse({
                'success': True,
                'message': 'Kategori berhasil ditambahkan',
                'data': {
                    'id': category.id,
                    'name': category.name,
                    'description': category.description
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_category_update(request, pk):
    """Update beneficiary category"""
    category = get_object_or_404(BeneficiaryCategory, pk=pk)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            category.name = data.get('name', category.name)
            category.description = data.get('description', category.description)
            category.criteria = data.get('criteria', category.criteria)
            category.is_active = data.get('is_active', category.is_active)
            category.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Kategori berhasil diperbarui',
                'data': {
                    'id': category.id,
                    'name': category.name,
                    'description': category.description
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    # GET request - return HTML form
    context = {
        'title': 'Edit Kategori Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'categories',
        'category': category,
    }
    return render(request, 'admin_panel/beneficiaries/category_form.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["DELETE"])
def admin_category_delete(request, pk):
    """Delete beneficiary category"""
    try:
        category = get_object_or_404(BeneficiaryCategory, pk=pk)
        
        # Check if category has beneficiaries
        if category.beneficiaries.exists():
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus kategori yang masih memiliki penerima bantuan'
            })
        
        name = category.name
        category.delete()
        return JsonResponse({
            'success': True,
            'message': f'Kategori {name} berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== AID PROGRAM MANAGEMENT ====================

@login_required
@user_passes_test(is_admin)
def admin_aid_programs_list(request):
    """Admin list view for aid programs"""
    aid_programs = Aid.objects.annotate(
        distribution_count=Count('distributions')
    ).order_by('-created_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status == 'active':
        aid_programs = aid_programs.filter(is_active=True)
    elif status == 'inactive':
        aid_programs = aid_programs.filter(is_active=False)
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        aid_programs = aid_programs.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(source__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(aid_programs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Manajemen Program Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'aid_programs',
        'aid_programs': page_obj,
        'search': search,
        'status_filter': status,
    }
    return render(request, 'admin_panel/beneficiaries/aid_programs.html', context)

@login_required
@user_passes_test(is_admin)
def admin_aid_program_create(request):
    """Create new aid program"""
    from .forms import AidForm
    
    if request.method == 'POST':
        form = AidForm(request.POST)
        if form.is_valid():
            aid_program = form.save(commit=False)
            aid_program.created_by = request.user
            aid_program.save()
            messages.success(request, 'Program bantuan berhasil ditambahkan')
            return redirect('admin_panel:beneficiaries_aid_programs')
        else:
            messages.error(request, 'Gagal menambahkan program bantuan. Periksa kembali data Anda.')
    else:
        form = AidForm()
    
        context = {
            'title': 'Tambah Program Bantuan',
        'form': form,
            'active_menu': 'beneficiaries',
            'active_submenu': 'aid_programs',
        }
        return render(request, 'admin_panel/beneficiaries/aid_program_form.html', context)

@login_required
@user_passes_test(is_admin)
def admin_aid_program_update(request, pk):
    """Update aid program"""
    from .forms import AidForm
    
    aid_program = get_object_or_404(Aid, pk=pk)
    
    if request.method == 'POST':
        form = AidForm(request.POST, instance=aid_program)
        if form.is_valid():
            form.save()
            messages.success(request, 'Program bantuan berhasil diperbarui')
            return redirect('admin_panel:beneficiaries_aid_programs')
        else:
            messages.error(request, 'Gagal memperbarui program bantuan. Periksa kembali data Anda.')
    else:
        form = AidForm(instance=aid_program)
    
    context = {
        'title': 'Edit Program Bantuan',
        'form': form,
        'aid_program': aid_program,
        'active_menu': 'beneficiaries',
        'active_submenu': 'aid_programs',
    }
    return render(request, 'admin_panel/beneficiaries/aid_program_form.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["DELETE"])
def admin_aid_program_delete(request, pk):
    """Delete aid program"""
    try:
        aid_program = get_object_or_404(Aid, pk=pk)
        
        # Check if program has distributions
        if aid_program.distributions.exists():
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus program yang sudah memiliki distribusi'
            })
        
        name = aid_program.name
        aid_program.delete()
        return JsonResponse({
            'success': True,
            'message': f'Program bantuan {name} berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== DISTRIBUTION MANAGEMENT ====================

@login_required
@user_passes_test(is_admin)
def admin_distributions_list(request):
    """Admin list view for aid distributions"""
    distributions = AidDistribution.objects.select_related(
        'beneficiary__person', 'aid'
    ).order_by('-created_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        distributions = distributions.filter(status=status)
    
    # Filter by aid program
    aid_id = request.GET.get('aid')
    if aid_id:
        distributions = distributions.filter(aid_id=aid_id)
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        distributions = distributions.filter(distribution_date__gte=date_from)
    if date_to:
        distributions = distributions.filter(distribution_date__lte=date_to)
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        distributions = distributions.filter(
            Q(beneficiary__person__name__icontains=search) |
            Q(beneficiary__person__nik__icontains=search) |
            Q(aid__name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(distributions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get aid programs for filter
    aid_programs = Aid.objects.filter(is_active=True)
    
    # Get aids for filter dropdown
    aids = Aid.objects.filter(is_active=True)
    
    context = {
        'title': 'Manajemen Distribusi Bantuan',
        'distributions': page_obj,
        'aids': aids,
        'selected_status': status,
        'active_menu': 'beneficiaries',
        'active_submenu': 'distributions',
        'selected_aid': aid_id,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
    }
    return render(request, 'admin_panel/beneficiaries/aid_distributions.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_distribution_create(request):
    """Create new aid distribution"""
    if request.method == 'GET':
        # Return HTML form for GET request
        from .forms import AidDistributionForm
        form = AidDistributionForm()
        beneficiaries = Beneficiary.objects.filter(status='aktif').select_related('person')
        aid_programs = Aid.objects.filter(is_active=True)
        
        context = {
            'title': 'Tambah Distribusi Bantuan',
            'active_menu': 'beneficiaries',
            'active_submenu': 'distributions_add',
            'form': form,
            'beneficiaries': beneficiaries,
            'aids': aid_programs,  # Changed from aid_programs to aids to match template
            'status_choices': AidDistribution.STATUS_CHOICES,
        }
        return render(request, 'admin_panel/beneficiaries/aid_distribution_form.html', context)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            distribution = AidDistribution.objects.create(
                beneficiary_id=data.get('beneficiary_id'),
                aid_id=data.get('aid_id'),
                amount_received=data.get('amount_received'),
                distribution_date=data.get('distribution_date'),
                notes=data.get('notes', ''),
                status=data.get('status', 'pending')
            )
            return JsonResponse({
                'success': True,
                'message': 'Distribusi bantuan berhasil ditambahkan',
                'data': {
                    'id': distribution.id,
                    'beneficiary': distribution.beneficiary.person.name,
                    'aid': distribution.aid.name,
                    'amount_received': str(distribution.amount_received)
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def admin_distribution_update_status(request, pk):
    """Update distribution status"""
    try:
        distribution = get_object_or_404(AidDistribution, pk=pk)
        data = json.loads(request.body)
        
        distribution.status = data.get('status', distribution.status)
        distribution.notes = data.get('notes', distribution.notes)
        
        if distribution.status == 'distributed' and not distribution.distribution_date:
            distribution.distribution_date = timezone.now().date()
        
        distribution.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Status distribusi berhasil diperbarui',
            'data': {
                'id': distribution.id,
                'status': distribution.status,
                'distribution_date': distribution.distribution_date.isoformat() if distribution.distribution_date else None
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== VERIFICATION MANAGEMENT ====================

@login_required
@user_passes_test(is_admin)
def admin_verifications_list(request):
    """Admin list view for beneficiary verifications"""
    verifications = BeneficiaryVerification.objects.select_related(
        'beneficiary__person', 'beneficiary__category', 'verifier'
    ).order_by('-created_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        verifications = verifications.filter(verification_status=status)
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        verifications = verifications.filter(
            Q(beneficiary__person__name__icontains=search) |
            Q(beneficiary__person__nik__icontains=search) |
            Q(verification_notes__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(verifications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    pending_count = BeneficiaryVerification.objects.filter(verification_status='pending').count()
    verified_count = BeneficiaryVerification.objects.filter(verification_status='verified').count()
    rejected_count = BeneficiaryVerification.objects.filter(verification_status='rejected').count()
    total_count = BeneficiaryVerification.objects.count()
    
    context = {
        'title': 'Manajemen Verifikasi Penerima Bantuan',
        'active_menu': 'beneficiaries',
        'active_submenu': 'verifications',
        'verifications': page_obj,
        'search': search,
        'status_filter': status,
        'pending_count': pending_count,
        'verified_count': verified_count,
        'rejected_count': rejected_count,
        'total_count': total_count,
    }
    return render(request, 'admin_panel/beneficiaries/verifications.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def admin_verification_update(request, pk):
    """Update verification status"""
    try:
        verification = get_object_or_404(BeneficiaryVerification, pk=pk)
        data = json.loads(request.body)
        
        action = data.get('action')
        notes = data.get('notes', '')
        documents_checked = data.get('documents_checked', '')
        field_visit_conducted = data.get('field_visit_conducted', False)
        field_visit_notes = data.get('field_visit_notes', '')
        
        # Update verification fields
        verification.verification_notes = notes
        verification.documents_checked = documents_checked
        verification.field_visit_conducted = field_visit_conducted
        verification.field_visit_notes = field_visit_notes
        verification.verifier = request.user
        verification.updated_at = timezone.now()
        
        # Set status based on action
        if action == 'approve':
            verification.verification_status = 'verified'
            message = 'Verifikasi berhasil di-approve'
        elif action == 'reject':
            verification.verification_status = 'rejected'
            message = 'Verifikasi berhasil di-reject'
        elif action == 'edit':
            verification.verification_status = data.get('status', verification.verification_status)
            message = 'Verifikasi berhasil diperbarui'
        
        verification.save()
        
        return JsonResponse({
            'success': True,
            'message': message,
            'data': {
                'id': verification.id,
                'status': verification.verification_status,
                'updated_at': verification.updated_at.isoformat()
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@user_passes_test(is_admin)
def admin_verification_detail(request, pk):
    """Get verification detail for modal"""
    try:
        verification = get_object_or_404(BeneficiaryVerification.objects.select_related(
            'beneficiary__person', 'beneficiary__category', 'verifier'
        ), pk=pk)
        
        html = f"""
        <div class="row">
            <div class="col-md-6">
                <h6>Informasi Penerima</h6>
                <table class="table table-sm">
                    <tr><td><strong>Nama:</strong></td><td>{verification.beneficiary.person.name}</td></tr>
                    <tr><td><strong>NIK:</strong></td><td>{verification.beneficiary.person.nik}</td></tr>
                    <tr><td><strong>Kategori:</strong></td><td>{verification.beneficiary.category.name}</td></tr>
                    <tr><td><strong>Status Ekonomi:</strong></td><td>{verification.beneficiary.get_economic_status_display()}</td></tr>
                </table>
            </div>
            <div class="col-md-6">
                <h6>Informasi Verifikasi</h6>
                <table class="table table-sm">
                    <tr><td><strong>Status:</strong></td><td><span class="badge bg-{'success' if verification.verification_status == 'verified' else 'warning' if verification.verification_status == 'pending' else 'danger'}">{verification.get_verification_status_display()}</span></td></tr>
                    <tr><td><strong>Tanggal:</strong></td><td>{verification.verification_date.strftime('%d/%m/%Y')}</td></tr>
                    <tr><td><strong>Verifikator:</strong></td><td>{verification.verifier.get_full_name() if verification.verifier else '-'}</td></tr>
                    <tr><td><strong>Kunjungan Lapangan:</strong></td><td>{'Ya' if verification.field_visit_conducted else 'Tidak'}</td></tr>
                </table>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6>Catatan Verifikasi</h6>
                <div class="border p-3 rounded">
                    {verification.verification_notes or 'Tidak ada catatan'}
                </div>
            </div>
        </div>
        
        {f'''
        <div class="row mt-3">
            <div class="col-12">
                <h6>Dokumen yang Diperiksa</h6>
                <div class="border p-3 rounded">
                    {verification.documents_checked or 'Tidak ada informasi dokumen'}
                </div>
            </div>
        </div>
        ''' if verification.documents_checked else ''}
        
        {f'''
        <div class="row mt-3">
            <div class="col-12">
                <h6>Catatan Kunjungan Lapangan</h6>
                <div class="border p-3 rounded">
                    {verification.field_visit_notes}
                </div>
            </div>
        </div>
        ''' if verification.field_visit_notes else ''}
        """
        
        return JsonResponse({
            'success': True,
            'html': html
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def admin_verification_bulk_action(request):
    """Bulk action for verifications"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        ids = data.get('ids', [])
        
        if not ids:
            return JsonResponse({
                'success': False,
                'message': 'Tidak ada verifikasi yang dipilih'
            })
        
        verifications = BeneficiaryVerification.objects.filter(id__in=ids)
        updated_count = 0
        
        for verification in verifications:
            verification.verifier = request.user
            verification.updated_at = timezone.now()
            
            if action == 'approve':
                verification.verification_status = 'verified'
                verification.verification_notes = 'Verifikasi massal - Approved'
            elif action == 'reject':
                verification.verification_status = 'rejected'
                verification.verification_notes = 'Verifikasi massal - Rejected'
            
            verification.save()
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} verifikasi berhasil diperbarui',
            'updated_count': updated_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== REPORTS AND ANALYTICS ====================

@login_required
@user_passes_test(is_admin)
def admin_beneficiaries_reports(request):
    """Generate comprehensive reports for beneficiaries with filters"""
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    import json
    
    # Get filter parameters
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    economic_filter = request.GET.get('economic_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    beneficiaries_qs = Beneficiary.objects.select_related('person', 'category')
    distributions_qs = AidDistribution.objects.select_related('beneficiary', 'aid')
    
    # Apply filters
    if category_filter:
        beneficiaries_qs = beneficiaries_qs.filter(category_id=category_filter)
        distributions_qs = distributions_qs.filter(beneficiary__category_id=category_filter)
    
    if status_filter:
        beneficiaries_qs = beneficiaries_qs.filter(status=status_filter)
    
    if economic_filter:
        beneficiaries_qs = beneficiaries_qs.filter(economic_status=economic_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__gte=date_from_obj)
            distributions_qs = distributions_qs.filter(created_at__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__lte=date_to_obj)
            distributions_qs = distributions_qs.filter(created_at__lte=date_to_obj)
        except:
            pass
    
    # Basic statistics
    total_beneficiaries = beneficiaries_qs.count()
    active_beneficiaries = beneficiaries_qs.filter(status='aktif').count()
    inactive_beneficiaries = beneficiaries_qs.filter(status='tidak_aktif').count()
    verified_beneficiaries = BeneficiaryVerification.objects.filter(
        beneficiary__in=beneficiaries_qs,
        verification_status='verified'
    ).count()
    
    # Category breakdown
    category_stats = BeneficiaryCategory.objects.annotate(
        count=Count('beneficiary', filter=Q(beneficiary__in=beneficiaries_qs)),
        active_count=Count('beneficiary', filter=Q(beneficiary__in=beneficiaries_qs, beneficiary__status='aktif')),
    ).filter(count__gt=0).order_by('-count')
    
    # Economic status breakdown
    economic_stats = beneficiaries_qs.values('economic_status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Status breakdown
    status_stats = beneficiaries_qs.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly registration trends (last 12 months)
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_stats = beneficiaries_qs.filter(
        created_at__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Prepare chart data
    month_labels = []
    month_values = []
    for stat in monthly_stats:
        if stat['month']:
            month_labels.append(stat['month'].strftime('%b %Y'))
            month_values.append(stat['count'])
    
    # Distribution statistics
    total_distributions = distributions_qs.count()
    pending_distributions = distributions_qs.filter(status='pending').count()
    approved_distributions = distributions_qs.filter(status='approved').count()
    completed_distributions = distributions_qs.filter(status='distributed').count()
    rejected_distributions = distributions_qs.filter(status='rejected').count()
    
    total_amount_distributed = distributions_qs.filter(
        status='distributed'
    ).aggregate(total=Sum('amount_received'))['total'] or 0
    
    # Distribution by aid program
    aid_distribution_stats = distributions_qs.filter(status='distributed').values(
        'aid__name', 'aid__aid_type'
    ).annotate(
        count=Count('id'),
        total_amount=Sum('amount_received')
    ).order_by('-total_amount')[:10]
    
    # Monthly distribution trends
    monthly_distribution_stats = distributions_qs.filter(
        created_at__gte=twelve_months_ago,
        status='distributed'
    ).annotate(
        month=TruncMonth('distribution_date')
    ).values('month').annotate(
        count=Count('id'),
        total_amount=Sum('amount_received')
    ).order_by('month')
    
    dist_month_labels = []
    dist_month_counts = []
    dist_month_amounts = []
    for stat in monthly_distribution_stats:
        if stat['month']:
            dist_month_labels.append(stat['month'].strftime('%b %Y'))
            dist_month_counts.append(stat['count'])
            dist_month_amounts.append(float(stat['total_amount'] or 0))
    
    # Get beneficiaries for table (paginated)
    beneficiaries_list = beneficiaries_qs.order_by('-created_at')[:100]
    
    # Categories for filter
    all_categories = BeneficiaryCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'title': 'Laporan Penerima Bantuan',
        'total_beneficiaries': total_beneficiaries,
        'active_beneficiaries': active_beneficiaries,
        'inactive_beneficiaries': inactive_beneficiaries,
        'verified_beneficiaries': verified_beneficiaries,
        'category_stats': category_stats,
        'economic_stats': economic_stats,
        'status_stats': status_stats,
        'monthly_stats': monthly_stats,
        'total_distributions': total_distributions,
        'pending_distributions': pending_distributions,
        'approved_distributions': approved_distributions,
        'completed_distributions': completed_distributions,
        'rejected_distributions': rejected_distributions,
        'total_amount_distributed': total_amount_distributed,
        'aid_distribution_stats': aid_distribution_stats,
        'beneficiaries_list': beneficiaries_list,
        'all_categories': all_categories,
        # Chart data
        'month_labels': json.dumps(month_labels),
        'month_values': json.dumps(month_values),
        'dist_month_labels': json.dumps(dist_month_labels),
        'dist_month_counts': json.dumps(dist_month_counts),
        'dist_month_amounts': json.dumps(dist_month_amounts),
        # Filter values
        'category_filter': category_filter,
        'status_filter': status_filter,
        'economic_filter': economic_filter,
        'date_from': date_from,
        'date_to': date_to,
        # Menu context untuk sidebar
        'active_menu': 'beneficiaries',
        'active_submenu': 'reports',
    }
    return render(request, 'admin_panel/beneficiaries/reports.html', context)

@login_required
@user_passes_test(is_admin)
def admin_export_beneficiaries_excel(request):
    """Export beneficiaries report to Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Get filter parameters (same as report view)
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    economic_filter = request.GET.get('economic_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply filters
    beneficiaries_qs = Beneficiary.objects.select_related('person', 'category')
    
    if category_filter:
        beneficiaries_qs = beneficiaries_qs.filter(category_id=category_filter)
    if status_filter:
        beneficiaries_qs = beneficiaries_qs.filter(status=status_filter)
    if economic_filter:
        beneficiaries_qs = beneficiaries_qs.filter(economic_status=economic_filter)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__gte=date_from_obj)
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__lte=date_to_obj)
        except:
            pass
    
    beneficiaries = beneficiaries_qs.order_by('-created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Penerima Bantuan'
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = 'LAPORAN DATA PENERIMA BANTUAN SOSIAL'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:L2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f'Desa Pulosarok - Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = [
        'No', 'NIK', 'Nama Lengkap', 'Jenis Kelamin', 'Kategori Bantuan',
        'Status Ekonomi', 'Pendapatan/Bulan', 'Jumlah Keluarga',
        'Kondisi Rumah', 'Status', 'Tanggal Daftar', 'Catatan'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 5
    for idx, b in enumerate(beneficiaries, 1):
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=b.person.nik).border = border
        ws.cell(row=row_num, column=3, value=b.person.name).border = border
        ws.cell(row=row_num, column=4, value=b.person.get_gender_display()).border = border
        ws.cell(row=row_num, column=5, value=b.category.name if b.category else '').border = border
        ws.cell(row=row_num, column=6, value=b.get_economic_status_display()).border = border
        ws.cell(row=row_num, column=7, value=float(b.monthly_income) if b.monthly_income else 0).border = border
        ws.cell(row=row_num, column=8, value=b.family_members_count).border = border
        ws.cell(row=row_num, column=9, value=b.house_condition[:50] if b.house_condition else '').border = border
        ws.cell(row=row_num, column=10, value=b.get_status_display()).border = border
        ws.cell(row=row_num, column=11, value=b.registration_date.strftime('%d-%m-%Y') if b.registration_date else '').border = border
        ws.cell(row=row_num, column=12, value=b.notes[:50] if b.notes else '').border = border
        row_num += 1
    
    # Adjust column widths
    column_widths = [5, 18, 25, 15, 20, 18, 15, 12, 30, 12, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Laporan_Penerima_Bantuan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
@user_passes_test(is_admin)
def admin_export_beneficiaries_pdf(request):
    """Export beneficiaries report to PDF"""
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime
    
    # Get filter parameters
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    economic_filter = request.GET.get('economic_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply filters
    beneficiaries_qs = Beneficiary.objects.select_related('person', 'category')
    
    if category_filter:
        beneficiaries_qs = beneficiaries_qs.filter(category_id=category_filter)
    if status_filter:
        beneficiaries_qs = beneficiaries_qs.filter(status=status_filter)
    if economic_filter:
        beneficiaries_qs = beneficiaries_qs.filter(economic_status=economic_filter)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__gte=date_from_obj)
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            beneficiaries_qs = beneficiaries_qs.filter(registration_date__lte=date_to_obj)
        except:
            pass
    
    beneficiaries = beneficiaries_qs.order_by('-created_at')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Laporan_Penerima_Bantuan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=18)
    
    # Container for elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Title
    title = Paragraph("LAPORAN DATA PENERIMA BANTUAN SOSIAL", title_style)
    elements.append(title)
    
    subtitle = Paragraph(f"Desa Pulosarok<br/>Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.2*inch))
    
    # Statistics summary
    stats_data = [
        ['Total Penerima', str(beneficiaries.count())],
        ['Aktif', str(beneficiaries.filter(status='aktif').count())],
        ['Tidak Aktif', str(beneficiaries.filter(status='tidak_aktif').count())],
    ]
    stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f9ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Data table
    data = [['No', 'NIK', 'Nama', 'Kategori', 'Status Ekonomi', 'Pendapatan', 'Status']]
    
    for idx, b in enumerate(beneficiaries[:50], 1):  # Limit 50 for PDF
        data.append([
            str(idx),
            b.person.nik[:12] + '...' if len(b.person.nik) > 12 else b.person.nik,
            b.person.name[:20] if len(b.person.name) > 20 else b.person.name,
            (b.category.name[:15] + '...') if b.category and len(b.category.name) > 15 else (b.category.name if b.category else ''),
            b.get_economic_status_display()[:15],
            f"Rp {int(b.monthly_income):,}" if b.monthly_income else 'Rp 0',
            b.get_status_display()
        ])
    
    # Create table
    table = Table(data, colWidths=[0.4*inch, 1.2*inch, 1.5*inch, 1.3*inch, 1.2*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Data
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    elements.append(table)
    
    # Footer note
    if beneficiaries.count() > 50:
        elements.append(Spacer(1, 0.2*inch))
        note_style = ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        note = Paragraph(f"<i>Catatan: Menampilkan 50 dari {beneficiaries.count()} total penerima bantuan</i>", note_style)
        elements.append(note)
    
    # Build PDF
    doc.build(elements)
    
    return response

# ==================== BULK OPERATIONS ====================

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def admin_bulk_update_status(request):
    """Bulk update beneficiary status"""
    try:
        data = json.loads(request.body)
        beneficiary_ids = data.get('ids', [])
        new_status = data.get('status')
        
        if not beneficiary_ids or not new_status:
            return JsonResponse({
                'success': False,
                'message': 'ID dan status harus disediakan'
            })
        
        updated_count = Beneficiary.objects.filter(
            id__in=beneficiary_ids
        ).update(status=new_status)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} penerima bantuan berhasil diperbarui',
            'updated_count': updated_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def admin_bulk_verify(request):
    """Bulk verify beneficiaries"""
    try:
        data = json.loads(request.body)
        beneficiary_ids = data.get('ids', [])
        
        if not beneficiary_ids:
            return JsonResponse({
                'success': False,
                'message': 'ID penerima bantuan harus disediakan'
            })
        
        # Update verification status - create verification records instead
        updated_count = 0
        
        # Create verification records
        beneficiaries = Beneficiary.objects.filter(id__in=beneficiary_ids)
        verifications = []
        for beneficiary in beneficiaries:
            verifications.append(BeneficiaryVerification(
                beneficiary=beneficiary,
                status='approved',
                notes='Verifikasi massal oleh admin',
                verified_by=request.user,
                verified_at=timezone.now()
            ))
        
        BeneficiaryVerification.objects.bulk_create(verifications)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} penerima bantuan berhasil diverifikasi',
            'verified_count': updated_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

# ==================== AJAX API ENDPOINTS ====================

@login_required
@user_passes_test(is_admin)
def api_beneficiary_search(request):
    """AJAX endpoint for beneficiary search"""
    query = request.GET.get('q', '').strip()
    if len(query) < 3:
        return JsonResponse({'success': True, 'data': []})
    
    beneficiaries = Beneficiary.objects.filter(
        Q(person__name__icontains=query) |
        Q(person__nik__icontains=query)
    ).select_related('person')[:10]
    
    results = []
    for b in beneficiaries:
        results.append({
            'id': b.id,
            'text': f'{b.person.name} - {b.person.nik}',
            'name': b.person.name,
            'nik': b.person.nik
        })
    
        return JsonResponse({'success': True, 'data': results})

@login_required
@user_passes_test(is_admin)
def api_beneficiary_stats(request):
    """AJAX endpoint for beneficiary statistics"""
    stats = {
        'total': Beneficiary.objects.count(),
        'active': Beneficiary.objects.filter(status='aktif').count(),
        'verified': BeneficiaryVerification.objects.filter(verification_status='verified').count(),
        'pending_verification': BeneficiaryVerification.objects.filter(verification_status='pending').count(),
        'categories': list(BeneficiaryCategory.objects.annotate(
            count=Count('beneficiary')
        ).values('name', 'count')),
        'economic_status': list(Beneficiary.objects.values('economic_status').annotate(
            count=Count('id')
        ).values('economic_status', 'count'))
    }
    return JsonResponse(stats)


def api_aid_stats(request):
    """API untuk statistik bantuan"""
    stats = {
        'total_aids': Aid.objects.filter(is_active=True).count(),
        'total_budget': float(Aid.objects.filter(
            is_active=True
        ).aggregate(total=Sum('total_budget'))['total'] or 0),
        'by_type': list(Aid.objects.filter(
            is_active=True
        ).values('aid_type').annotate(
            count=Count('id'),
            total_budget=Sum('total_budget')
        ).values('aid_type', 'count', 'total_budget')),
        'distributions': {
            'distributed': AidDistribution.objects.filter(status='distributed').count(),
            'pending': AidDistribution.objects.filter(status='pending').count(),
            'approved': AidDistribution.objects.filter(status='approved').count()
        }
    }
    
    return JsonResponse(stats)

def admin_distribution_detail(request, pk):
    """Admin distribution detail view"""
    try:
        distribution = AidDistribution.objects.get(pk=pk)
        context = {
            'active_menu': 'beneficiaries',
            'active_submenu': 'distributions',
            'distribution': distribution,
        }
        return render(request, 'admin_panel/beneficiaries/aid_distribution_detail.html', context)
    except AidDistribution.DoesNotExist:
        messages.error(request, 'Distribusi bantuan tidak ditemukan')
        return redirect('admin_panel:aid_distributions_list')

def admin_distribution_update(request, pk):
    """Admin distribution update view"""
    try:
        distribution = AidDistribution.objects.get(pk=pk)
        if request.method == 'POST':
            # Handle update logic here
            pass
        context = {
            'active_menu': 'beneficiaries',
            'active_submenu': 'distributions',
            'distribution': distribution,
        }
        return render(request, 'admin_panel/beneficiaries/aid_distribution_form.html', context)
    except AidDistribution.DoesNotExist:
        messages.error(request, 'Distribusi bantuan tidak ditemukan')
        return redirect('admin_panel:aid_distributions_list')

def aid_distribution_check(request):
    """Check if aid distribution exists"""
    aid_id = request.GET.get('aid_id')
    beneficiary_id = request.GET.get('beneficiary_id')
    receipt_number = request.GET.get('receipt_number')
    exclude_id = request.GET.get('exclude_id')
    
    exists = False
    
    if aid_id and beneficiary_id:
        queryset = AidDistribution.objects.filter(aid_id=aid_id, beneficiary_id=beneficiary_id)
        if exclude_id:
            queryset = queryset.exclude(id=exclude_id)
        exists = queryset.exists()
    
    if receipt_number:
        queryset = AidDistribution.objects.filter(receipt_number=receipt_number)
        if exclude_id:
            queryset = queryset.exclude(id=exclude_id)
        exists = queryset.exists()
    
    return JsonResponse({'exists': exists})

def admin_category_detail(request, pk):
    """Admin category detail view"""
    try:
        category = BeneficiaryCategory.objects.get(pk=pk)
        beneficiaries = Beneficiary.objects.filter(category=category)
        context = {
            'active_menu': 'beneficiaries',
            'active_submenu': 'categories',
            'category': category,
            'beneficiaries': beneficiaries,
        }
        return render(request, 'admin_panel/beneficiaries/category_detail.html', context)
    except BeneficiaryCategory.DoesNotExist:
        messages.error(request, 'Kategori tidak ditemukan')
        return redirect('admin_panel:categories_list')